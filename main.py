import streamlit as st
import sqlite3
import pandas as pd
from datetime import datetime, date
import hashlib
import bcrypt
from cryptography.fernet import Fernet
from PIL import Image
import io
import json
import plotly.express as px
import plotly.graph_objects as go
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import qrcode
from io import BytesIO
import re
import base64
import os
import time
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Kumasi District YoKA Registration System",
    page_icon="⛪",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==================== ENCRYPTION SETUP ====================
def get_encryption_key():
    key_file = Path("encryption.key")
    if key_file.exists():
        with open(key_file, "rb") as f:
            return f.read()
    else:
        key = Fernet.generate_key()
        with open(key_file, "wb") as f:
            f.write(key)
        return key

ENCRYPTION_KEY = get_encryption_key()
cipher = Fernet(ENCRYPTION_KEY)

def encrypt_password(password):
    if not password:
        return ""
    return cipher.encrypt(password.encode()).decode()

def decrypt_password(encrypted_password):
    if not encrypted_password or encrypted_password is None:
        return ""
    try:
        return cipher.decrypt(encrypted_password.encode()).decode()
    except Exception:
        return ""

# ==================== VALIDATION FUNCTIONS ====================
def validate_ghana_phone(phone):
    if not phone:
        return False
    phone_str = str(phone)
    phone_str = re.sub(r'[\s\-+]', '', phone_str)
    patterns = [
        r'^0[2-5]\d{8}$',
        r'^233[2-5]\d{8}$',
        r'^\+233[2-5]\d{8}$'
    ]
    for pattern in patterns:
        if re.match(pattern, phone_str):
            return True
    return False

def normalize_phone(phone):
    if not phone:
        return ""
    phone_str = str(phone)
    phone_str = re.sub(r'[\s\-]', '', phone_str)
    if phone_str.startswith('+233'):
        phone_str = '0' + phone_str[4:]
    elif phone_str.startswith('233'):
        phone_str = '0' + phone_str[3:]
    return phone_str

def calculate_age_from_dob(dob):
    if not dob:
        return 18
    today = date.today()
    if isinstance(dob, datetime):
        dob = dob.date()
    return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))

def validate_email(email):
    if not email:
        return True
    pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    return re.match(pattern, email) is not None

# ==================== DATABASE SETUP ====================
def init_rbac_database():
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()

    c.execute('''CREATE TABLE IF NOT EXISTS branches
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  branch_name TEXT UNIQUE NOT NULL,
                  branch_code TEXT UNIQUE NOT NULL,
                  location TEXT,
                  contact_person TEXT,
                  contact_phone TEXT,
                  created_date TEXT NOT NULL,
                  is_active BOOLEAN DEFAULT 1)''')

    c.execute('''CREATE TABLE IF NOT EXISTS users
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  username TEXT UNIQUE NOT NULL,
                  password TEXT NOT NULL,
                  full_name TEXT NOT NULL,
                  email TEXT,
                  phone TEXT,
                  role TEXT NOT NULL CHECK(role IN ('super_admin', 'admin', 'branch_executive')),
                  assigned_branch_id INTEGER,
                  is_active BOOLEAN DEFAULT 1,
                  created_by INTEGER,
                  created_date TEXT NOT NULL,
                  last_login TEXT,
                  FOREIGN KEY (assigned_branch_id) REFERENCES branches(id),
                  FOREIGN KEY (created_by) REFERENCES users(id))''')

    c.execute('''CREATE TABLE IF NOT EXISTS members
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  official_name TEXT NOT NULL,
                  date_of_birth TEXT,
                  age INTEGER NOT NULL,
                  residence TEXT NOT NULL,
                  active_phone TEXT NOT NULL,
                  email TEXT,
                  profile_picture BLOB,
                  school_name TEXT NOT NULL,
                  school_level TEXT NOT NULL,
                  school_class TEXT NOT NULL,
                  school_house TEXT,
                  residence_status TEXT NOT NULL,
                  residence_name TEXT,
                  church_branch TEXT NOT NULL,
                  branch_id INTEGER NOT NULL,
                  yoka_hall TEXT NOT NULL,
                  youth_camps_attended INTEGER NOT NULL,
                  has_church_position BOOLEAN DEFAULT 0,
                  church_position_status TEXT,
                  church_position_type TEXT,
                  church_position_name TEXT,
                  church_position_duration TEXT,
                  work_status TEXT,
                  work_type TEXT,
                  work_name TEXT,
                  work_position TEXT,
                  work_location TEXT,
                  work_experience_years INTEGER,
                  is_diaspora BOOLEAN DEFAULT 0,
                  diaspora_country TEXT,
                  diaspora_status TEXT,
                  diaspora_job TEXT,
                  diaspora_school TEXT,
                  diaspora_education_level TEXT,
                  mother_name TEXT,
                  mother_phone TEXT,
                  mother_occupation TEXT,
                  father_name TEXT,
                  father_phone TEXT,
                  father_occupation TEXT,
                  guardian_name TEXT,
                  guardian_phone TEXT,
                  guardian_relationship TEXT,
                  guardian_occupation TEXT,
                  submission_date TEXT NOT NULL,
                  created_by INTEGER,
                  last_modified_by INTEGER,
                  last_modified_date TEXT,
                  is_verified BOOLEAN DEFAULT 0,
                  verified_by INTEGER,
                  verified_date TEXT,
                  gender TEXT,
                  emergency_contact_name TEXT,
                  emergency_contact_phone TEXT,
                  medical_conditions TEXT,
                  talents TEXT,
                  interests TEXT,
                  programme TEXT,
                  courses TEXT,
                  hall_or_hostel TEXT,
                  form TEXT,
                  FOREIGN KEY (branch_id) REFERENCES branches(id),
                  FOREIGN KEY (created_by) REFERENCES users(id))''')

    c.execute('''CREATE TABLE IF NOT EXISTS audit_log
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  user_id INTEGER,
                  username TEXT,
                  action TEXT NOT NULL,
                  entity_type TEXT NOT NULL,
                  entity_id INTEGER,
                  details TEXT,
                  ip_address TEXT,
                  timestamp TEXT NOT NULL,
                  FOREIGN KEY (user_id) REFERENCES users(id))''')

    c.execute('''CREATE TABLE IF NOT EXISTS system_settings
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  setting_key TEXT UNIQUE NOT NULL,
                  setting_value TEXT,
                  setting_type TEXT DEFAULT 'text',
                  description TEXT)''')

    c.execute('''CREATE TABLE IF NOT EXISTS email_settings
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  smtp_server TEXT,
                  smtp_port INTEGER,
                  sender_email TEXT,
                  sender_password TEXT,
                  use_tls BOOLEAN DEFAULT 1,
                  updated_by INTEGER,
                  updated_date TEXT)''')

    c.execute('''CREATE TABLE IF NOT EXISTS email_logs
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  recipient_email TEXT,
                  subject TEXT,
                  message TEXT,
                  status TEXT,
                  sent_date TEXT,
                  sent_by INTEGER)''')

    c.execute("CREATE INDEX IF NOT EXISTS idx_members_branch ON members(branch_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_members_phone ON members(active_phone)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_members_date ON members(submission_date)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_members_name ON members(official_name)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_members_school ON members(school_name)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_users_role ON users(role)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_users_branch ON users(assigned_branch_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_audit_timestamp ON audit_log(timestamp)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_audit_user ON audit_log(user_id)")

    default_branches = [
        ('Kumasi Central', 'KMC', 'Kumasi Central Business District', 'Rev. John Doe', '0244000001'),
        ('Asokwa', 'ASK', 'Asokwa District', 'Rev. Jane Smith', '0244000002'),
        ('Tafo', 'TAF', 'Tafo Community', 'Rev. Michael Brown', '0244000003'),
        ('Suame', 'SUA', 'Suame Magazine Area', 'Rev. Sarah Johnson', '0244000004'),
        ('Bantama', 'BAN', 'Bantama Community', 'Rev. David Wilson', '0244000005')
    ]
    for branch in default_branches:
        c.execute("INSERT OR IGNORE INTO branches (branch_name, branch_code, location, contact_person, contact_phone, created_date) VALUES (?, ?, ?, ?, ?, ?)",
                  (branch[0], branch[1], branch[2], branch[3], branch[4], datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

    c.execute("SELECT * FROM users WHERE username='super_admin'")
    if not c.fetchone():
        hashed_password = bcrypt.hashpw('SuperAdmin@123'.encode(), bcrypt.gensalt()).decode()
        c.execute("INSERT INTO users (username, password, full_name, email, phone, role, created_date) VALUES (?, ?, ?, ?, ?, ?, ?)",
                  ('super_admin', hashed_password, 'Super Administrator', 'superadmin@yoka.org', '0244000000', 'super_admin', datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

    c.execute("SELECT * FROM users WHERE username='admin'")
    if not c.fetchone():
        hashed_password = bcrypt.hashpw('Admin@123'.encode(), bcrypt.gensalt()).decode()
        c.execute("INSERT INTO users (username, password, full_name, email, phone, role, assigned_branch_id, created_date) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                  ('admin', hashed_password, 'District Administrator', 'admin@yoka.org', '0244000006', 'admin', 1, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

    branch_execs = [
        ('exec_central', 'Central@123', 'Central Branch Executive', 'central@yoka.org', '0244000010', 1),
        ('exec_asokwa', 'Asokwa@123', 'Asokwa Branch Executive', 'asokwa@yoka.org', '0244000011', 2),
        ('exec_tafo', 'Tafo@123', 'Tafo Branch Executive', 'tafo@yoka.org', '0244000012', 3)
    ]
    for exec_data in branch_execs:
        c.execute("SELECT * FROM users WHERE username=?", (exec_data[0],))
        if not c.fetchone():
            hashed_password = bcrypt.hashpw(exec_data[1].encode(), bcrypt.gensalt()).decode()
            c.execute("INSERT INTO users (username, password, full_name, email, phone, role, assigned_branch_id, created_date) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                      (exec_data[0], hashed_password, exec_data[2], exec_data[3], exec_data[4], 'branch_executive', exec_data[5], datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

    default_settings = [
        ('system_name', 'Kumasi District YoKA Registration System', 'text', 'System display name'),
        ('primary_color', '#667eea', 'color', 'Primary theme color'),
        ('secondary_color', '#764ba2', 'color', 'Secondary theme color'),
        ('accent_color', '#f093fb', 'color', 'Accent theme color'),
        ('background_color', '#f8f9fa', 'color', 'Background color'),
        ('sidebar_color', '#1a1a2e', 'color', 'Sidebar color'),
        ('font_family', 'Inter', 'text', 'Font family'),
        ('font_size', '16px', 'text', 'Base font size'),
        ('card_border_radius', '12', 'text', 'Card border radius'),
        ('logo_url', '', 'text', 'URL to system logo'),
        ('favicon', '⛪', 'text', 'Favicon emoji or URL'),
        ('footer_text', '© 2024 Kumasi District YoKA. All Rights Reserved.', 'text', 'Footer text'),
        ('registration_open', 'true', 'boolean', 'Whether registration is open')
    ]
    for setting in default_settings:
        c.execute("INSERT OR IGNORE INTO system_settings (setting_key, setting_value, setting_type, description) VALUES (?, ?, ?, ?)",
                  (setting[0], setting[1], setting[2], setting[3]))

    conn.commit()
    conn.close()

# ==================== SYSTEM SETTINGS FUNCTIONS ====================
def get_system_settings():
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    df = pd.read_sql_query("SELECT * FROM system_settings", conn)
    conn.close()
    return df

def update_system_setting(setting_key, setting_value, updated_by):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()
    c.execute("UPDATE system_settings SET setting_value = ? WHERE setting_key = ?", (setting_value, setting_key))
    conn.commit()
    conn.close()
    log_audit(updated_by, st.session_state.get('username', 'system'), 'UPDATE_SETTING', 'system', None, f"Updated setting: {setting_key} = {setting_value}")
    return True

def apply_custom_styling():
    st.markdown("""
    <style>
        /* FORCE LIGHT MODE - Override Streamlit's dark mode detection */
        html, body, [data-testid="stAppViewContainer"], .stApp {
            background-color: #ffffff !important;
        }
        
        [data-testid="stHeader"] {
            background-color: #ffffff !important;
        }
        
        [data-testid="stToolbar"] {
            background-color: #ffffff !important;
        }
        
        /* Main content area */
        .main > div {
            background-color: #ffffff !important;
        }
        
        /* Block containers */
        .stMarkdown, .stTextInput, .stSelectbox, .stTextArea, .stDateInput {
            background-color: #ffffff !important;
        }
        
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
        * { font-family: 'Inter', sans-serif; }
        .main .block-container {
            padding-top: 1rem;
            padding-bottom: 0rem;
            max-width: 1400px;
        }
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
        header {visibility: hidden;}
        ::-webkit-scrollbar { width: 8px; height: 8px; }
        ::-webkit-scrollbar-track { background: #f1f1f1; border-radius: 10px; }
        ::-webkit-scrollbar-thumb {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            border-radius: 10px;
        }
        [data-testid="stSidebar"] {
            background: linear-gradient(180deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
            box-shadow: 2px 0 10px rgba(0,0,0,0.1);
        }
        [data-testid="stSidebar"] * { color: #ffffff !important; }
        [data-testid="stSidebar"] .stMarkdown { color: #ffffff !important; }
        [data-testid="stSidebar"] h1,
        [data-testid="stSidebar"] h2,
        [data-testid="stSidebar"] h3,
        [data-testid="stSidebar"] h4,
        [data-testid="stSidebar"] p,
        [data-testid="stSidebar"] span,
        [data-testid="stSidebar"] div,
        [data-testid="stSidebar"] label { color: #ffffff !important; }
        .stButton > button {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            border-radius: 10px;
            padding: 0.6rem 1.2rem;
            font-weight: 600;
            transition: all 0.3s ease;
            width: 100%;
            box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        }
        .stButton > button:hover {
            transform: translateY(-2px);
            box-shadow: 0 5px 15px rgba(102,126,234,0.4);
            background: linear-gradient(135deg, #764ba2 0%, #667eea 100%);
        }
        .stForm {
            background: white;
            padding: 1.5rem;
            border-radius: 20px;
            box-shadow: 0 4px 15px rgba(0,0,0,0.08);
            margin-bottom: 1rem;
        }
        .stTextInput > div > div > input,
        .stTextArea > div > div > textarea,
        .stSelectbox > div > div > select,
        .stDateInput > div > div > input {
            border-radius: 10px;
            border: 2px solid #e0e0e0;
            transition: all 0.3s ease;
            background-color: #ffffff !important;
            color: #000000 !important;
        }
        .stTextInput > div > div > input:focus,
        .stTextArea > div > div > textarea:focus {
            border-color: #667eea;
            box-shadow: 0 0 0 2px rgba(102,126,234,0.2);
        }
        h1, h2, h3, h4, h5, h6 { color: #1a1a2e; font-weight: 600; }
        [data-testid="stMetric"] {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 1rem;
            border-radius: 15px;
            color: white;
            box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        }
        .stDataFrame { border-radius: 15px; overflow: hidden; }
        .stTabs [data-baseweb="tab-list"] {
            gap: 8px;
            background: #f5f5f5;
            padding: 8px;
            border-radius: 12px;
        }
        .stTabs [data-baseweb="tab"] {
            border-radius: 8px;
            padding: 8px 16px;
            transition: all 0.3s ease;
        }
        .stTabs [aria-selected="true"] {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white !important;
        }
        .stAlert {
            border-radius: 12px;
            border-left: 4px solid;
            animation: slideIn 0.5s ease-out;
        }
        @keyframes slideIn {
            from { opacity: 0; transform: translateX(-20px); }
            to { opacity: 1; transform: translateX(0); }
        }
        .streamlit-expanderHeader {
            background: rgba(102,126,234,0.1);
            border-radius: 10px;
            font-weight: 600;
        }
        .badge {
            display: inline-block;
            padding: 4px 8px;
            border-radius: 12px;
            font-size: 12px;
            font-weight: 600;
        }
        .badge-success { background: #d4edda; color: #155724; }
        .badge-warning { background: #fff3cd; color: #856404; }
        @media (max-width: 768px) {
            .main .block-container { padding: 1rem; }
        }
        
        /* Force text colors for light mode */
        p, span, div, label {
            color: #000000 !important;
        }
        
        /* Override any dark mode specific elements */
        @media (prefers-color-scheme: dark) {
            html, body, .stApp, [data-testid="stAppViewContainer"] {
                background-color: #ffffff !important;
            }
            p, span, div, label {
                color: #000000 !important;
            }
        }
    </style>
    """, unsafe_allow_html=True)

# ==================== EMAIL FUNCTIONS ====================
def get_email_settings():
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()
    c.execute("SELECT * FROM email_settings ORDER BY id DESC LIMIT 1")
    settings = c.fetchone()
    conn.close()

    if settings:
        try:
            smtp_server = settings[1] if len(settings) > 1 else None
            smtp_port = settings[2] if len(settings) > 2 else None
            sender_email = settings[3] if len(settings) > 3 else None
            encrypted_password = settings[4] if len(settings) > 4 else None
            use_tls = settings[5] if len(settings) > 5 else 1
            decrypted_password = decrypt_password(encrypted_password) if encrypted_password else ""
        except Exception as e:
            print(f"Error decrypting email settings: {e}")
            decrypted_password = ""
            smtp_server = None
            smtp_port = 587
            sender_email = None
            use_tls = True

        return {
            'smtp_server': smtp_server,
            'smtp_port': smtp_port or 587,
            'sender_email': sender_email,
            'sender_password': decrypted_password,
            'use_tls': bool(use_tls) if use_tls is not None else True
        }
    return None

def save_email_settings(smtp_server, smtp_port, sender_email, sender_password, use_tls, updated_by):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()
    c.execute("DELETE FROM email_settings")
    encrypted_password = encrypt_password(sender_password) if sender_password else ""
    c.execute("""INSERT INTO email_settings
                 (smtp_server, smtp_port, sender_email, sender_password, use_tls, updated_by, updated_date)
                 VALUES (?, ?, ?, ?, ?, ?, ?)""",
              (smtp_server, smtp_port, sender_email, encrypted_password, use_tls, updated_by, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit()
    conn.close()

def send_email(recipient_email, subject, body):
    if not recipient_email:
        return False, "No recipient email provided"

    settings = get_email_settings()
    if not settings or not settings['smtp_server'] or not settings['sender_email']:
        return False, "Email settings not configured"

    try:
        msg = MIMEMultipart()
        msg['From'] = settings['sender_email']
        msg['To'] = recipient_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'html'))

        server = smtplib.SMTP(settings['smtp_server'], settings['smtp_port'])
        if settings['use_tls']:
            server.starttls()
        server.login(settings['sender_email'], settings['sender_password'])
        server.send_message(msg)
        server.quit()

        conn = sqlite3.connect('kumasi_yoka_registration.db')
        c = conn.cursor()
        c.execute("""INSERT INTO email_logs (recipient_email, subject, message, status, sent_date, sent_by)
                     VALUES (?, ?, ?, ?, ?, ?)""",
                  (recipient_email, subject, body, 'sent', datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                   st.session_state.get('user_id', 1)))
        conn.commit()
        conn.close()

        return True, "Email sent successfully"
    except Exception as e:
        return False, str(e)

def send_user_credentials_email(email, username, password, full_name, role):
    if not email:
        return False, "No email address provided"

    subject = "Welcome to YoKA Registration System - Your Login Credentials"
    body = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; }}
            .container {{ padding: 20px; background-color: #f4f4f4; }}
            .header {{ background: linear-gradient(135deg, #667eea, #764ba2); color: white; padding: 10px; text-align: center; }}
            .content {{ background-color: white; padding: 20px; margin: 20px 0; }}
            .credentials {{ background-color: #f0f0f0; padding: 15px; margin: 10px 0; }}
            .footer {{ text-align: center; font-size: 12px; color: #666; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header"><h2>Kumasi District YoKA Registration System</h2></div>
            <div class="content">
                <h3>Dear {full_name},</h3>
                <p>Your account has been created in the YoKA Registration System with the following credentials:</p>
                <div class="credentials">
                    <p><strong>Username:</strong> {username}</p>
                    <p><strong>Password:</strong> {password}</p>
                    <p><strong>Role:</strong> {role.replace('_', ' ').title()}</p>
                </div>
                <p><strong>Important:</strong> Please change your password after your first login.</p>
                <p>You can access the system using the login page.</p>
                <h4>Security Tips:</h4>
                <ul>
                    <li>Never share your password with anyone</li>
                    <li>Use a strong password</li>
                    <li>Log out after each session</li>
                </ul>
            </div>
            <div class="footer">
                <p>This is an automated message, please do not reply.</p>
                <p>© 2024 Kumasi District YoKA. All Rights Reserved.</p>
            </div>
        </div>
    </body>
    </html>
    """
    return send_email(email, subject, body)

# ==================== HELPER FUNCTIONS ====================
def hash_password(password):
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password, hashed):
    return bcrypt.checkpw(password.encode(), hashed.encode())

def log_audit(user_id, username, action, entity_type, entity_id, details=""):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()
    c.execute("INSERT INTO audit_log (user_id, username, action, entity_type, entity_id, details, timestamp) VALUES (?, ?, ?, ?, ?, ?, ?)",
              (user_id, username, action, entity_type, entity_id, details, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit()
    conn.close()

def get_user_info(username):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()
    c.execute("""SELECT u.id, u.username, u.full_name, u.email, u.phone, u.role, u.assigned_branch_id, u.is_active,
                        b.branch_name, b.branch_code
                 FROM users u
                 LEFT JOIN branches b ON u.assigned_branch_id = b.id
                 WHERE u.username = ?""", (username,))
    user = c.fetchone()
    conn.close()

    if user:
        return {
            'id': user[0],
            'username': user[1],
            'full_name': user[2],
            'email': user[3],
            'phone': user[4],
            'role': user[5],
            'assigned_branch_id': user[6],
            'is_active': user[7],
            'branch_name': user[8],
            'branch_code': user[9]
        }
    return None

def check_login(username, password):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()
    c.execute("SELECT * FROM users WHERE username=? AND is_active=1", (username,))
    user = c.fetchone()
    conn.close()

    if user and verify_password(password, user[2]):
        user_info = get_user_info(username)
        if user_info:
            conn = sqlite3.connect('kumasi_yoka_registration.db')
            c = conn.cursor()
            c.execute("UPDATE users SET last_login = ? WHERE id = ?", (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), user_info['id']))
            conn.commit()
            conn.close()
            log_audit(user_info['id'], username, 'LOGIN', 'user', user_info['id'], "User logged in")
            return user_info
    return None

# ==================== CRUD OPERATIONS FOR BRANCHES ====================
def create_branch(branch_name, branch_code, location, contact_person, contact_phone, created_by):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()
    try:
        c.execute("""INSERT INTO branches (branch_name, branch_code, location, contact_person, contact_phone, created_date)
                     VALUES (?, ?, ?, ?, ?, ?)""",
                  (branch_name, branch_code, location, contact_person, contact_phone, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        branch_id = c.lastrowid
        conn.commit()
        log_audit(created_by, st.session_state.get('username', 'system'), 'CREATE_BRANCH', 'branch', branch_id, f"Created branch: {branch_name}")
        return True, branch_id
    except sqlite3.IntegrityError:
        return False, "Branch name or code already exists"
    finally:
        conn.close()

def get_all_branches(include_inactive=False):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    if include_inactive:
        df = pd.read_sql_query("SELECT * FROM branches ORDER BY branch_name", conn)
    else:
        df = pd.read_sql_query("SELECT * FROM branches WHERE is_active = 1 ORDER BY branch_name", conn)
    conn.close()
    return df

def get_branch_by_id(branch_id):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()
    c.execute("SELECT * FROM branches WHERE id = ?", (branch_id,))
    branch = c.fetchone()
    conn.close()
    return branch

def update_branch(branch_id, branch_name, branch_code, location, contact_person, contact_phone, updated_by):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()
    c.execute("""UPDATE branches
                 SET branch_name = ?, branch_code = ?, location = ?, contact_person = ?, contact_phone = ?
                 WHERE id = ?""",
              (branch_name, branch_code, location, contact_person, contact_phone, branch_id))
    conn.commit()
    log_audit(updated_by, st.session_state.get('username', 'system'), 'UPDATE_BRANCH', 'branch', branch_id, f"Updated branch: {branch_name}")
    conn.close()
    return True

def delete_branch(branch_id, deleted_by):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()
    c.execute("UPDATE branches SET is_active = 0 WHERE id = ?", (branch_id,))
    conn.commit()
    log_audit(deleted_by, st.session_state.get('username', 'system'), 'DELETE_BRANCH', 'branch', branch_id, f"Deactivated branch ID: {branch_id}")
    conn.close()
    return True

# ==================== CRUD OPERATIONS FOR USERS ====================
def create_user(username, password, full_name, email, phone, role, assigned_branch_id, created_by):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()
    try:
        if phone and not validate_ghana_phone(phone):
            return False, "Invalid Ghana phone number format"

        hashed_password = hash_password(password)
        c.execute("""INSERT INTO users (username, password, full_name, email, phone, role, assigned_branch_id, created_by, created_date)
                     VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                  (username, hashed_password, full_name, email, phone, role, assigned_branch_id, created_by, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        user_id = c.lastrowid
        conn.commit()
        log_audit(created_by, st.session_state.get('username', 'system'), 'CREATE_USER', 'user', user_id, f"Created user: {username} with role: {role}")

        if email:
            send_user_credentials_email(email, username, password, full_name, role)

        return True, user_id
    except sqlite3.IntegrityError:
        return False, "Username already exists"
    finally:
        conn.close()

def reset_user_password(user_id, new_password, updated_by):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()
    hashed_password = hash_password(new_password)

    c.execute("SELECT email, full_name, username FROM users WHERE id = ?", (user_id,))
    user = c.fetchone()

    c.execute("UPDATE users SET password = ? WHERE id = ?", (hashed_password, user_id))
    conn.commit()

    if user and user[0]:
        subject = "YoKA System - Password Reset Notification"
        body = f"""
        <html>
        <body>
            <h3>Password Reset</h3>
            <p>Dear {user[1]},</p>
            <p>Your password has been reset. Your new login credentials are:</p>
            <p><strong>Username:</strong> {user[2]}</p>
            <p><strong>New Password:</strong> {new_password}</p>
            <p>Please change your password after logging in.</p>
        </body>
        </html>
        """
        send_email(user[0], subject, body)

    log_audit(updated_by, st.session_state.get('username', 'system'), 'RESET_PASSWORD', 'user', user_id, f"Reset password for user ID: {user_id}")
    conn.close()
    return True

def get_all_users():
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    query = """SELECT u.id, u.username, u.full_name, u.email, u.phone, u.role,
                      b.branch_name as assigned_branch, u.is_active, u.created_date, u.last_login
               FROM users u
               LEFT JOIN branches b ON u.assigned_branch_id = b.id
               ORDER BY u.created_date DESC"""
    df = pd.read_sql_query(query, conn)
    conn.close()
    return df

def get_user_by_id(user_id):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()
    c.execute("SELECT * FROM users WHERE id = ?", (user_id,))
    user = c.fetchone()
    conn.close()
    return user

def update_user(user_id, full_name, email, phone, role, assigned_branch_id, is_active, updated_by):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()

    if phone and not validate_ghana_phone(phone):
        conn.close()
        return False, "Invalid Ghana phone number format"

    c.execute("""UPDATE users
                 SET full_name = ?, email = ?, phone = ?, role = ?, assigned_branch_id = ?, is_active = ?
                 WHERE id = ?""",
              (full_name, email, phone, role, assigned_branch_id, is_active, user_id))
    conn.commit()
    log_audit(updated_by, st.session_state.get('username', 'system'), 'UPDATE_USER', 'user', user_id, f"Updated user ID: {user_id}")
    conn.close()
    return True, "User updated successfully"

def delete_user(user_id, deleted_by):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()
    c.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    log_audit(deleted_by, st.session_state.get('username', 'system'), 'DELETE_USER', 'user', user_id, f"Deleted user ID: {user_id}")
    conn.close()
    return True

# ==================== CRUD OPERATIONS FOR MEMBERS ====================
def save_member(data, profile_picture=None, created_by=None):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()

    if not validate_ghana_phone(data['active_phone']):
        conn.close()
        raise ValueError("Invalid Ghana phone number format")

    if data.get('email') and not validate_email(data['email']):
        conn.close()
        raise ValueError("Invalid email format")

    c.execute("SELECT id FROM branches WHERE branch_name = ?", (data['church_branch'],))
    branch_result = c.fetchone()
    branch_id = branch_result[0] if branch_result else 1

    profile_pic_bytes = None
    if profile_picture:
        img = Image.open(profile_picture)
        img.thumbnail((300, 300))
        buffer = BytesIO()
        img.save(buffer, format="JPEG", quality=85)
        profile_pic_bytes = buffer.getvalue()

    c.execute('''INSERT INTO members (
            official_name, date_of_birth, age, residence, active_phone, email, profile_picture,
            school_name, school_level, school_class, school_house, residence_status, residence_name,
            church_branch, branch_id, yoka_hall, youth_camps_attended,
            has_church_position, church_position_status, church_position_type, church_position_name, church_position_duration,
            work_status, work_type, work_name, work_position, work_location, work_experience_years,
            is_diaspora, diaspora_country, diaspora_status, diaspora_job, diaspora_school, diaspora_education_level,
            mother_name, mother_phone, mother_occupation, father_name, father_phone, father_occupation,
            guardian_name, guardian_phone, guardian_relationship, guardian_occupation,
            submission_date, created_by, last_modified_by, last_modified_date,
            is_verified, verified_by, verified_date, gender, emergency_contact_name,
            emergency_contact_phone, medical_conditions, talents, interests,
            programme, courses, hall_or_hostel, form
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
              (data['official_name'], data.get('date_of_birth', ''), data['age'], data['residence'],
               normalize_phone(data['active_phone']), data.get('email', ''), profile_pic_bytes,
               data['school_name'], data['school_level'], data.get('school_class', ''),
               data.get('school_house', ''), data['residence_status'], data.get('residence_name', ''),
               data['church_branch'], branch_id, data['yoka_hall'], data['youth_camps_attended'],
               1 if data.get('has_church_position') else 0, data.get('church_position_status', ''), data.get('church_position_type', ''),
               data.get('church_position_name', ''), data.get('church_position_duration', ''),
               data.get('work_status', ''), data.get('work_type', ''), data.get('work_name', ''), data.get('work_position', ''),
               data.get('work_location', ''), data.get('work_experience_years', 0),
               1 if data.get('is_diaspora') else 0, data.get('diaspora_country', ''), data.get('diaspora_status', ''), data.get('diaspora_job', ''),
               data.get('diaspora_school', ''), data.get('diaspora_education_level', ''),
               data.get('mother_name', ''), data.get('mother_phone', ''), data.get('mother_occupation', ''),
               data.get('father_name', ''), data.get('father_phone', ''), data.get('father_occupation', ''),
               data.get('guardian_name', ''), data.get('guardian_phone', ''), data.get('guardian_relationship', ''), data.get('guardian_occupation', ''),
               datetime.now().strftime("%Y-%m-%d %H:%M:%S"), created_by, created_by, datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
               0, None, None,
               data.get('gender', ''), data.get('emergency_contact_name', ''),
               normalize_phone(data.get('emergency_contact_phone', '')) if data.get('emergency_contact_phone') else '',
               data.get('medical_conditions', ''), data.get('talents', ''), data.get('interests', ''),
               data.get('programme', ''), data.get('courses', ''), data.get('hall_or_hostel', ''), data.get('form', '')))

    member_id = c.lastrowid
    conn.commit()
    log_audit(created_by, st.session_state.get('username', 'system'), 'CREATE_MEMBER', 'member', member_id, f"Created member: {data['official_name']}")
    conn.close()
    return member_id

def get_members_by_role(user_role, user_branch_id=None, user_id=None):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    try:
        if user_role in ('super_admin', 'admin'):
            df = pd.read_sql_query("SELECT * FROM members ORDER BY submission_date DESC", conn)
        else:
            df = pd.read_sql_query("SELECT * FROM members WHERE branch_id = ? ORDER BY submission_date DESC", conn, params=(user_branch_id,))
    except Exception as e:
        print(f"Error in get_members_by_role: {e}")
        df = pd.DataFrame()
    finally:
        conn.close()
    return df

def get_member_by_id(member_id, user_role, user_branch_id=None):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()
    try:
        if user_role == 'branch_executive':
            c.execute("SELECT * FROM members WHERE id = ? AND branch_id = ?", (member_id, user_branch_id))
        else:
            c.execute("SELECT * FROM members WHERE id = ?", (member_id,))
        member = c.fetchone()
    except Exception as e:
        print(f"Error in get_member_by_id: {e}")
        member = None
    finally:
        conn.close()
    return member

def update_member(member_id, data, updated_by, profile_picture=None):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()

    if not validate_ghana_phone(data['active_phone']):
        conn.close()
        raise ValueError("Invalid Ghana phone number format")

    if data.get('email') and not validate_email(data['email']):
        conn.close()
        raise ValueError("Invalid email format")

    c.execute("SELECT id FROM branches WHERE branch_name = ?", (data['church_branch'],))
    branch_result = c.fetchone()
    branch_id = branch_result[0] if branch_result else 1

    if profile_picture:
        img = Image.open(profile_picture)
        img.thumbnail((300, 300))
        buffer = BytesIO()
        img.save(buffer, format="JPEG", quality=85)
        profile_pic_bytes = buffer.getvalue()

        c.execute('''UPDATE members SET
                 official_name = ?, date_of_birth = ?, age = ?, residence = ?, active_phone = ?, email = ?,
                 profile_picture = ?,
                 school_name = ?, school_level = ?, school_class = ?, school_house = ?,
                 residence_status = ?, residence_name = ?, church_branch = ?, branch_id = ?,
                 yoka_hall = ?, youth_camps_attended = ?,
                 has_church_position = ?, church_position_status = ?, church_position_type = ?,
                 church_position_name = ?, church_position_duration = ?,
                 work_status = ?, work_type = ?, work_name = ?, work_position = ?,
                 work_location = ?, work_experience_years = ?,
                 is_diaspora = ?, diaspora_country = ?, diaspora_status = ?, diaspora_job = ?,
                 diaspora_school = ?, diaspora_education_level = ?,
                 mother_name = ?, mother_phone = ?, mother_occupation = ?,
                 father_name = ?, father_phone = ?, father_occupation = ?,
                 guardian_name = ?, guardian_phone = ?, guardian_relationship = ?, guardian_occupation = ?,
                 last_modified_by = ?, last_modified_date = ?, gender = ?, emergency_contact_name = ?,
                 emergency_contact_phone = ?, medical_conditions = ?, talents = ?, interests = ?,
                 programme = ?, courses = ?, hall_or_hostel = ?, form = ?
                 WHERE id = ?''',
              (data['official_name'], data.get('date_of_birth', ''), data['age'], data['residence'],
               normalize_phone(data['active_phone']), data.get('email', ''), profile_pic_bytes,
               data['school_name'], data['school_level'], data.get('school_class', ''),
               data.get('school_house', ''),
               data['residence_status'], data.get('residence_name', ''), data['church_branch'], branch_id,
               data['yoka_hall'], data['youth_camps_attended'],
               1 if data.get('has_church_position') else 0, data.get('church_position_status', ''), data.get('church_position_type', ''),
               data.get('church_position_name', ''), data.get('church_position_duration', ''),
               data.get('work_status', ''), data.get('work_type', ''), data.get('work_name', ''), data.get('work_position', ''),
               data.get('work_location', ''), data.get('work_experience_years', 0),
               1 if data.get('is_diaspora') else 0, data.get('diaspora_country', ''), data.get('diaspora_status', ''), data.get('diaspora_job', ''),
               data.get('diaspora_school', ''), data.get('diaspora_education_level', ''),
               data.get('mother_name', ''), normalize_phone(data.get('mother_phone', '')) if data.get('mother_phone') else '', data.get('mother_occupation', ''),
               data.get('father_name', ''), normalize_phone(data.get('father_phone', '')) if data.get('father_phone') else '', data.get('father_occupation', ''),
               data.get('guardian_name', ''), normalize_phone(data.get('guardian_phone', '')) if data.get('guardian_phone') else '', data.get('guardian_relationship', ''), data.get('guardian_occupation', ''),
               updated_by, datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
               data.get('gender', ''), data.get('emergency_contact_name', ''),
               normalize_phone(data.get('emergency_contact_phone', '')) if data.get('emergency_contact_phone') else '',
               data.get('medical_conditions', ''), data.get('talents', ''), data.get('interests', ''),
               data.get('programme', ''), data.get('courses', ''), data.get('hall_or_hostel', ''), data.get('form', ''), member_id))
    else:
        c.execute('''UPDATE members SET
                 official_name = ?, date_of_birth = ?, age = ?, residence = ?, active_phone = ?, email = ?,
                 school_name = ?, school_level = ?, school_class = ?, school_house = ?,
                 residence_status = ?, residence_name = ?, church_branch = ?, branch_id = ?,
                 yoka_hall = ?, youth_camps_attended = ?,
                 has_church_position = ?, church_position_status = ?, church_position_type = ?,
                 church_position_name = ?, church_position_duration = ?,
                 work_status = ?, work_type = ?, work_name = ?, work_position = ?,
                 work_location = ?, work_experience_years = ?,
                 is_diaspora = ?, diaspora_country = ?, diaspora_status = ?, diaspora_job = ?,
                 diaspora_school = ?, diaspora_education_level = ?,
                 mother_name = ?, mother_phone = ?, mother_occupation = ?,
                 father_name = ?, father_phone = ?, father_occupation = ?,
                 guardian_name = ?, guardian_phone = ?, guardian_relationship = ?, guardian_occupation = ?,
                 last_modified_by = ?, last_modified_date = ?, gender = ?, emergency_contact_name = ?,
                 emergency_contact_phone = ?, medical_conditions = ?, talents = ?, interests = ?,
                 programme = ?, courses = ?, hall_or_hostel = ?, form = ?
                 WHERE id = ?''',
              (data['official_name'], data.get('date_of_birth', ''), data['age'], data['residence'],
               normalize_phone(data['active_phone']), data.get('email', ''),
               data['school_name'], data['school_level'], data.get('school_class', ''),
               data.get('school_house', ''),
               data['residence_status'], data.get('residence_name', ''), data['church_branch'], branch_id,
               data['yoka_hall'], data['youth_camps_attended'],
               1 if data.get('has_church_position') else 0, data.get('church_position_status', ''), data.get('church_position_type', ''),
               data.get('church_position_name', ''), data.get('church_position_duration', ''),
               data.get('work_status', ''), data.get('work_type', ''), data.get('work_name', ''), data.get('work_position', ''),
               data.get('work_location', ''), data.get('work_experience_years', 0),
               1 if data.get('is_diaspora') else 0, data.get('diaspora_country', ''), data.get('diaspora_status', ''), data.get('diaspora_job', ''),
               data.get('diaspora_school', ''), data.get('diaspora_education_level', ''),
               data.get('mother_name', ''), normalize_phone(data.get('mother_phone', '')) if data.get('mother_phone') else '', data.get('mother_occupation', ''),
               data.get('father_name', ''), normalize_phone(data.get('father_phone', '')) if data.get('father_phone') else '', data.get('father_occupation', ''),
               data.get('guardian_name', ''), normalize_phone(data.get('guardian_phone', '')) if data.get('guardian_phone') else '', data.get('guardian_relationship', ''), data.get('guardian_occupation', ''),
               updated_by, datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
               data.get('gender', ''), data.get('emergency_contact_name', ''),
               normalize_phone(data.get('emergency_contact_phone', '')) if data.get('emergency_contact_phone') else '',
               data.get('medical_conditions', ''), data.get('talents', ''), data.get('interests', ''),
               data.get('programme', ''), data.get('courses', ''), data.get('hall_or_hostel', ''), data.get('form', ''), member_id))

    conn.commit()
    log_audit(updated_by, st.session_state.get('username', 'system'), 'UPDATE_MEMBER', 'member', member_id, f"Updated member ID: {member_id}")
    conn.close()
    return True

def delete_member(member_id, deleted_by, user_role, user_branch_id=None):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()

    if user_role == 'branch_executive':
        c.execute("SELECT branch_id FROM members WHERE id = ?", (member_id,))
        member = c.fetchone()
        if member and member[0] != user_branch_id:
            conn.close()
            return False, "Permission denied: Cannot delete members from other branches"

    c.execute("DELETE FROM members WHERE id = ?", (member_id,))
    conn.commit()
    log_audit(deleted_by, st.session_state.get('username', 'system'), 'DELETE_MEMBER', 'member', member_id, f"Deleted member ID: {member_id}")
    conn.close()
    return True, "Member deleted successfully"

def verify_member(member_id, verified_by):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()
    c.execute("""UPDATE members
                 SET is_verified = 1, verified_by = ?, verified_date = ?
                 WHERE id = ?""",
              (verified_by, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), member_id))
    conn.commit()
    log_audit(verified_by, st.session_state.get('username', 'system'), 'VERIFY_MEMBER', 'member', member_id, f"Verified member ID: {member_id}")
    conn.close()
    return True

def delete_bulk_members(member_ids, deleted_by):
    conn = sqlite3.connect('kumasi_yoka_registration.db')
    c = conn.cursor()
    placeholders = ','.join('?' * len(member_ids))
    c.execute(f"DELETE FROM members WHERE id IN ({placeholders})", member_ids)
    conn.commit()
    count = c.rowcount
    log_audit(deleted_by, st.session_state.get('username', 'system'), 'BULK_DELETE', 'members', count, f"Bulk deleted {count} members")
    conn.close()
    return count

# ==================== IMPORT/EXPORT FUNCTIONS ====================
def export_to_excel(df, include_photos=False):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        export_df = df.drop(columns=['profile_picture'], errors='ignore')
        export_df.to_excel(writer, sheet_name='Members', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Members']

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    return output.getvalue()

def import_from_csv(file, target_branch, created_by):
    df = pd.read_csv(file)
    success_count = 0
    errors = []

    for idx, row in df.iterrows():
        try:
            data = {
                'official_name': str(row.get('official_name', '')),
                'date_of_birth': str(row.get('date_of_birth', '')) if pd.notna(row.get('date_of_birth')) else '',
                'age': int(row.get('age', 18)) if pd.notna(row.get('age')) else 18,
                'residence': str(row.get('residence', '')),
                'active_phone': str(row.get('active_phone', '')),
                'email': str(row.get('email', '')),
                'school_name': str(row.get('school_name', '')),
                'school_level': str(row.get('school_level', 'SHS')),
                'school_class': str(row.get('school_class', '')),
                'school_house': str(row.get('school_house', '')),
                'residence_status': str(row.get('residence_status', 'Day Student')),
                'residence_name': str(row.get('residence_name', '')),
                'church_branch': target_branch,
                'yoka_hall': str(row.get('yoka_hall', 'Hall A')),
                'youth_camps_attended': int(row.get('youth_camps_attended', 0)) if pd.notna(row.get('youth_camps_attended')) else 0,
                'has_church_position': bool(row.get('has_church_position', False)),
                'church_position_status': str(row.get('church_position_status', '')),
                'church_position_type': str(row.get('church_position_type', '')),
                'church_position_name': str(row.get('church_position_name', '')),
                'church_position_duration': str(row.get('church_position_duration', '')),
                'work_status': str(row.get('work_status', '')),
                'work_type': str(row.get('work_type', '')),
                'work_name': str(row.get('work_name', '')),
                'work_position': str(row.get('work_position', '')),
                'work_location': str(row.get('work_location', '')),
                'work_experience_years': int(row.get('work_experience_years', 0)) if pd.notna(row.get('work_experience_years')) else 0,
                'is_diaspora': bool(row.get('is_diaspora', False)),
                'diaspora_country': str(row.get('diaspora_country', '')),
                'diaspora_status': str(row.get('diaspora_status', '')),
                'diaspora_job': str(row.get('diaspora_job', '')),
                'diaspora_school': str(row.get('diaspora_school', '')),
                'diaspora_education_level': str(row.get('diaspora_education_level', '')),
                'mother_name': str(row.get('mother_name', '')),
                'mother_phone': str(row.get('mother_phone', '')),
                'mother_occupation': str(row.get('mother_occupation', '')),
                'father_name': str(row.get('father_name', '')),
                'father_phone': str(row.get('father_phone', '')),
                'father_occupation': str(row.get('father_occupation', '')),
                'guardian_name': str(row.get('guardian_name', '')),
                'guardian_phone': str(row.get('guardian_phone', '')),
                'guardian_relationship': str(row.get('guardian_relationship', '')),
                'guardian_occupation': str(row.get('guardian_occupation', '')),
                'gender': str(row.get('gender', '')),
                'emergency_contact_name': str(row.get('emergency_contact_name', '')),
                'emergency_contact_phone': str(row.get('emergency_contact_phone', '')),
                'medical_conditions': str(row.get('medical_conditions', '')),
                'talents': str(row.get('talents', '')),
                'interests': str(row.get('interests', '')),
                'programme': str(row.get('programme', '')),
                'courses': str(row.get('courses', '')),
                'hall_or_hostel': str(row.get('hall_or_hostel', '')),
                'form': str(row.get('form', ''))
            }
            save_member(data, profile_picture=None, created_by=created_by)
            success_count += 1
        except Exception as e:
            errors.append(f"Row {idx+2}: {str(e)}")

    return success_count, errors

def import_from_excel(file, target_branch, created_by):
    df = pd.read_excel(file)
    success_count = 0
    errors = []

    for idx, row in df.iterrows():
        try:
            data = {
                'official_name': str(row.get('official_name', '')),
                'date_of_birth': str(row.get('date_of_birth', '')) if pd.notna(row.get('date_of_birth')) else '',
                'age': int(row.get('age', 18)) if pd.notna(row.get('age')) else 18,
                'residence': str(row.get('residence', '')),
                'active_phone': str(row.get('active_phone', '')),
                'email': str(row.get('email', '')),
                'school_name': str(row.get('school_name', '')),
                'school_level': str(row.get('school_level', 'SHS')),
                'school_class': str(row.get('school_class', '')),
                'school_house': str(row.get('school_house', '')),
                'residence_status': str(row.get('residence_status', 'Day Student')),
                'residence_name': str(row.get('residence_name', '')),
                'church_branch': target_branch,
                'yoka_hall': str(row.get('yoka_hall', 'Hall A')),
                'youth_camps_attended': int(row.get('youth_camps_attended', 0)) if pd.notna(row.get('youth_camps_attended')) else 0,
                'has_church_position': bool(row.get('has_church_position', False)),
                'church_position_status': str(row.get('church_position_status', '')),
                'church_position_type': str(row.get('church_position_type', '')),
                'church_position_name': str(row.get('church_position_name', '')),
                'church_position_duration': str(row.get('church_position_duration', '')),
                'work_status': str(row.get('work_status', '')),
                'work_type': str(row.get('work_type', '')),
                'work_name': str(row.get('work_name', '')),
                'work_position': str(row.get('work_position', '')),
                'work_location': str(row.get('work_location', '')),
                'work_experience_years': int(row.get('work_experience_years', 0)) if pd.notna(row.get('work_experience_years')) else 0,
                'is_diaspora': bool(row.get('is_diaspora', False)),
                'diaspora_country': str(row.get('diaspora_country', '')),
                'diaspora_status': str(row.get('diaspora_status', '')),
                'diaspora_job': str(row.get('diaspora_job', '')),
                'diaspora_school': str(row.get('diaspora_school', '')),
                'diaspora_education_level': str(row.get('diaspora_education_level', '')),
                'mother_name': str(row.get('mother_name', '')),
                'mother_phone': str(row.get('mother_phone', '')),
                'mother_occupation': str(row.get('mother_occupation', '')),
                'father_name': str(row.get('father_name', '')),
                'father_phone': str(row.get('father_phone', '')),
                'father_occupation': str(row.get('father_occupation', '')),
                'guardian_name': str(row.get('guardian_name', '')),
                'guardian_phone': str(row.get('guardian_phone', '')),
                'guardian_relationship': str(row.get('guardian_relationship', '')),
                'guardian_occupation': str(row.get('guardian_occupation', '')),
                'gender': str(row.get('gender', '')),
                'emergency_contact_name': str(row.get('emergency_contact_name', '')),
                'emergency_contact_phone': str(row.get('emergency_contact_phone', '')),
                'medical_conditions': str(row.get('medical_conditions', '')),
                'talents': str(row.get('talents', '')),
                'interests': str(row.get('interests', '')),
                'programme': str(row.get('programme', '')),
                'courses': str(row.get('courses', '')),
                'hall_or_hostel': str(row.get('hall_or_hostel', '')),
                'form': str(row.get('form', ''))
            }
            save_member(data, profile_picture=None, created_by=created_by)
            success_count += 1
        except Exception as e:
            errors.append(f"Row {idx+2}: {str(e)}")

    return success_count, errors

# ==================== UI FUNCTIONS ====================
def advanced_export_ui():
    st.title("📥 Advanced Data Export")

    if st.session_state.user_role == 'branch_executive':
        members_df = get_members_by_role(st.session_state.user_role, st.session_state.user_branch_id)
        st.info(f"Exporting data from {st.session_state.branch_name} branch")
    else:
        members_df = get_members_by_role(st.session_state.user_role)
        branches_df = get_all_branches()
        if not branches_df.empty:
            branch_options = ["All Branches"] + branches_df['branch_name'].tolist()
            selected_branch = st.selectbox("Filter by Branch", branch_options)
            if selected_branch != "All Branches":
                members_df = members_df[members_df['church_branch'] == selected_branch]
                st.info(f"Exporting data from {selected_branch} branch")
            else:
                st.info("Exporting data from ALL branches")

    if not members_df.empty:
        st.subheader("Export Options")

        col1, col2 = st.columns(2)
        with col1:
            export_format = st.selectbox("Format", ["Excel (.xlsx)", "CSV", "JSON"])

        with col2:
            date_range = st.date_input("Date Range", [])
            if isinstance(date_range, (list, tuple)) and len(date_range) == 2:
                members_df['submission_date'] = pd.to_datetime(members_df['submission_date'])
                members_df = members_df[(members_df['submission_date'] >= pd.to_datetime(date_range[0])) &
                                       (members_df['submission_date'] <= pd.to_datetime(date_range[1]))]

        st.subheader("Select Columns to Export")
        exclude_columns = ['profile_picture', 'branch_id', 'created_by', 'last_modified_by']
        available_columns = [col for col in members_df.columns if col not in exclude_columns]

        selected_columns = st.multiselect("Columns", available_columns, default=available_columns[:15] if len(available_columns) > 15 else available_columns)

        if selected_columns:
            export_df = members_df[selected_columns]

            st.subheader("Preview")
            st.dataframe(export_df.head(10), use_container_width=True)

            st.info(f"📊 Total records to export: {len(export_df)}")

            if st.button("📥 Export Data", type="primary", use_container_width=True):
                if export_format == "Excel (.xlsx)":
                    excel_data = export_to_excel(export_df, include_photos=False)
                    st.download_button("Download Excel File", excel_data,
                                     file_name=f"yoka_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                elif export_format == "CSV":
                    csv = export_df.to_csv(index=False)
                    st.download_button("Download CSV", csv,
                                     file_name=f"yoka_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
                elif export_format == "JSON":
                    json_str = export_df.to_json(orient='records', indent=2)
                    st.download_button("Download JSON", json_str,
                                     file_name=f"yoka_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")

                log_audit(st.session_state.user_id, st.session_state.username, 'EXPORT_DATA',
                         'members', len(export_df), f"Exported {len(export_df)} records in {export_format}")
    else:
        st.info("No data to export")

def import_data_ui():
    st.title("📤 Import Data")

    if st.session_state.user_role not in ['super_admin', 'admin']:
        st.error("You don't have permission to import data. Only Super Admin and Admin can import data.")
        return

    branches_df = get_all_branches()
    if branches_df.empty:
        st.error("No branches available. Please create branches first.")
        return

    target_branch = st.selectbox("Import to Branch", branches_df['branch_name'].tolist())

    st.info("""
    **Import Instructions:**
    - File must contain column headers matching the database fields
    - Required fields: official_name, age, active_phone, school_name, at least one parent/guardian
    - Download the template below for correct format
    """)

    template_df = pd.DataFrame({
        'official_name': ['John Doe', 'Jane Smith'],
        'age': [18, 19],
        'active_phone': ['024XXXXXXX', '020XXXXXXX'],
        'school_name': ['Kumasi High School', 'Opoku Ware School'],
        'mother_name': ['Mrs. Jane Doe', 'Mrs. Mary Smith'],
        'mother_phone': ['024YYYYYYY', '020YYYYYYY'],
        'father_name': ['Mr. John Doe Sr.', 'Mr. James Smith'],
        'father_phone': ['024ZZZZZZZ', '020ZZZZZZZ'],
        'has_church_position': [True, False],
        'church_position_status': ['Present', ''],
        'church_position_type': ['YoKA Executive', ''],
        'church_position_duration': ['2 years', ''],
        'work_status': ['Internship', 'Permanent Staff'],
        'diaspora_status': ['Working', ''],
        'email': ['john@example.com', 'jane@example.com'],
        'residence': ['Kumasi', 'Asokwa'],
        'school_level': ['SHS', 'Tertiary'],
        'yoka_hall': ['Hall A', 'Hall B'],
        'gender': ['Male', 'Female']
    })

    col1, col2 = st.columns(2)
    with col1:
        st.download_button("📥 Download CSV Template", template_df.to_csv(index=False),
                         file_name="import_template.csv", mime="text/csv")
    with col2:
        excel_output = BytesIO()
        template_df.to_excel(excel_output, index=False)
        st.download_button("📥 Download Excel Template", excel_output.getvalue(),
                         file_name="import_template.xlsx")

    uploaded_file = st.file_uploader("Choose file to import", type=['csv', 'xlsx'])

    if uploaded_file:
        st.subheader("Preview of data to import")

        if uploaded_file.name.endswith('.csv'):
            uploaded_file.seek(0)
            df_preview = pd.read_csv(uploaded_file)
        else:
            uploaded_file.seek(0)
            df_preview = pd.read_excel(uploaded_file)

        st.dataframe(df_preview.head(10), use_container_width=True)
        st.write(f"Total records in file: {len(df_preview)}")

        col1, col2 = st.columns(2)
        with col1:
            if st.button("✅ Confirm Import", type="primary", use_container_width=True):
                with st.spinner("Importing data..."):
                    if uploaded_file.name.endswith('.csv'):
                        uploaded_file.seek(0)
                        success_count, errors = import_from_csv(uploaded_file, target_branch, st.session_state.user_id)
                    else:
                        uploaded_file.seek(0)
                        success_count, errors = import_from_excel(uploaded_file, target_branch, st.session_state.user_id)

                    if success_count > 0:
                        st.success(f"✅ Successfully imported {success_count} records to {target_branch} branch!")
                        log_audit(st.session_state.user_id, st.session_state.username, 'IMPORT_DATA',
                                 'members', success_count, f"Imported {success_count} members to {target_branch}")

                    if errors:
                        st.warning(f"⚠️ {len(errors)} errors occurred:")
                        for error in errors[:10]:
                            st.write(f"- {error}")
        with col2:
            if st.button("❌ Cancel", use_container_width=True):
                st.rerun()

def email_settings_ui():
    if st.session_state.user_role != 'super_admin':
        st.error("Only Super Admin can configure email settings")
        return

    st.header("📧 Email Configuration")

    current_settings = get_email_settings()

    with st.form("email_config_form"):
        col1, col2 = st.columns(2)
        with col1:
            smtp_server = st.text_input("SMTP Server",
                                       value=current_settings['smtp_server'] if current_settings and current_settings['smtp_server'] else "smtp.gmail.com")
            smtp_port = st.number_input("SMTP Port",
                                       value=int(current_settings['smtp_port']) if current_settings else 587)
            sender_email = st.text_input("Sender Email",
                                        value=current_settings['sender_email'] if current_settings and current_settings['sender_email'] else "")

        with col2:
            sender_password = st.text_input("Sender Password", type="password", value="")
            use_tls = st.checkbox("Use TLS", value=current_settings['use_tls'] if current_settings else True)
            test_email = st.text_input("Test Email Address (for sending test)")

        st.info("""
        **Email Configuration Tips:**
        - For Gmail: Use smtp.gmail.com, port 587, enable "Less secure app access" or use App Password
        - For Office 365: Use smtp.office365.com, port 587
        - For custom domains: Check with your hosting provider for SMTP settings
        """)

        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            submitted = st.form_submit_button("💾 Save Configuration", use_container_width=True)
        with col_btn2:
            test_email_btn = st.form_submit_button("📨 Send Test Email", use_container_width=True)

        if submitted:
            if sender_email and sender_password:
                save_email_settings(smtp_server, smtp_port, sender_email, sender_password, use_tls,
                                  st.session_state.user_id)
                st.success("Email settings saved successfully!")
            else:
                st.error("Please provide sender email and password")

        if test_email_btn:
            if test_email:
                success, message = send_email(test_email, "YoKA System - Test Email",
                                             "<h3>Test Email</h3><p>Your email configuration is working correctly!</p>")
                if success:
                    st.success("Test email sent successfully!")
                else:
                    st.error(f"Failed to send test email: {message}")
            else:
                st.warning("Please enter a test email address above.")

def theme_customizer_ui():
    if st.session_state.user_role != 'super_admin':
        st.error("Only Super Admin can customize themes")
        return

    st.title("🎨 Theme Customizer")

    settings_df = get_system_settings()
    settings = {row['setting_key']: row['setting_value'] for _, row in settings_df.iterrows()} if not settings_df.empty else {}

    col1, col2 = st.columns([1, 1])

    with col1:
        st.subheader("🎨 Color Scheme")
        primary_color = st.color_picker("Primary Color", settings.get('primary_color', '#667eea'))
        secondary_color = st.color_picker("Secondary Color", settings.get('secondary_color', '#764ba2'))
        accent_color = st.color_picker("Accent Color", settings.get('accent_color', '#f093fb'))
        background_color = st.color_picker("Background Color", settings.get('background_color', '#f8f9fa'))
        sidebar_color = st.color_picker("Sidebar Color", settings.get('sidebar_color', '#1a1a2e'))

        st.subheader("📝 Typography")
        font_options = ["Inter", "Poppins", "Roboto", "Arial", "Helvetica"]
        current_font = settings.get('font_family', 'Inter')
        font_index = font_options.index(current_font) if current_font in font_options else 0
        font_family = st.selectbox("Font Family", font_options, index=font_index)

        font_size = st.select_slider("Base Font Size",
                                     options=["12px", "14px", "16px", "18px", "20px"],
                                     value=settings.get('font_size', '16px'))

        st.subheader("📐 Layout")
        try:
            current_radius = int(settings.get('card_border_radius', '12'))
        except:
            current_radius = 12
        card_border_radius = st.slider("Card Border Radius", 0, 20, current_radius)

        st.subheader("🏷️ Branding")
        system_name = st.text_input("System Name", settings.get('system_name', 'Kumasi District YoKA Registration System'))
        logo_url = st.text_input("Logo URL (optional)", settings.get('logo_url', ''))
        favicon = st.text_input("Favicon Emoji", settings.get('favicon', '⛪'))
        footer_text = st.text_input("Footer Text", settings.get('footer_text', '© 2024 Kumasi District YoKA. All Rights Reserved.'))

        registration_open = st.checkbox("Registration Open", settings.get('registration_open', 'true') == 'true')

        if st.button("💾 Save Theme Settings", type="primary", use_container_width=True):
            update_system_setting('primary_color', primary_color, st.session_state.user_id)
            update_system_setting('secondary_color', secondary_color, st.session_state.user_id)
            update_system_setting('accent_color', accent_color, st.session_state.user_id)
            update_system_setting('background_color', background_color, st.session_state.user_id)
            update_system_setting('sidebar_color', sidebar_color, st.session_state.user_id)
            update_system_setting('font_family', font_family, st.session_state.user_id)
            update_system_setting('font_size', font_size, st.session_state.user_id)
            update_system_setting('card_border_radius', str(card_border_radius), st.session_state.user_id)
            update_system_setting('system_name', system_name, st.session_state.user_id)
            update_system_setting('logo_url', logo_url, st.session_state.user_id)
            update_system_setting('favicon', favicon, st.session_state.user_id)
            update_system_setting('footer_text', footer_text, st.session_state.user_id)
            update_system_setting('registration_open', str(registration_open).lower(), st.session_state.user_id)
            st.success("Theme saved successfully! Refreshing...")
            st.rerun()

    with col2:
        st.subheader("🔍 Live Preview")
        st.markdown(f"""
        <div style="background-color: {background_color}; padding: 20px; border-radius: {card_border_radius}px;">
            <h1 style="color: {primary_color};">{system_name}</h1>
            <div style="background-color: {sidebar_color}; padding: 15px; border-radius: {card_border_radius}px; margin: 10px 0;">
                <p style="color: white;">Sidebar navigation preview</p>
            </div>
            <button style="background: linear-gradient(135deg, {primary_color}, {secondary_color}); color: white; padding: 10px 20px; border: none; border-radius: 5px;">
                Button Preview
            </button>
            <div style="background: linear-gradient(135deg, {primary_color}, {secondary_color}); color: white; padding: 10px; margin-top: 10px; border-radius: {card_border_radius}px;">
                <p>Gradient preview</p>
            </div>
            <div style="margin-top: 10px; font-family: '{font_family}', sans-serif; font-size: {font_size};">
                <p>Font preview: The quick brown fox jumps over the lazy dog.</p>
            </div>
        </div>
        """, unsafe_allow_html=True)

def analytics_dashboard_ui():
    st.title("📊 YoKA Analytics Dashboard")

    if st.session_state.user_role == 'branch_executive':
        members_df = get_members_by_role(st.session_state.user_role, st.session_state.user_branch_id)
        st.info(f"📈 Analytics for {st.session_state.branch_name} Branch")
    else:
        members_df = get_members_by_role(st.session_state.user_role)
        branches_df = get_all_branches()
        if not branches_df.empty:
            branch_options = ["All Branches"] + branches_df['branch_name'].tolist()
            selected_branch = st.selectbox("Filter by Branch", branch_options)
            if selected_branch != "All Branches":
                members_df = members_df[members_df['church_branch'] == selected_branch]
                st.info(f"📈 Analytics for {selected_branch} Branch")
            else:
                st.info("📈 Analytics for ALL Branches")

    if not members_df.empty:
        members_df['submission_date'] = pd.to_datetime(members_df['submission_date'])

        st.subheader("📈 Key Performance Indicators")
        col1, col2, col3, col4, col5 = st.columns(5)

        with col1:
            total_members = len(members_df)
            new_this_month = len(members_df[members_df['submission_date'] > datetime.now() - pd.Timedelta(days=30)])
            st.metric("Total Members", total_members, delta=f"+{new_this_month} this month")

        with col2:
            avg_age = round(members_df['age'].mean(), 1) if len(members_df) > 0 else 0
            st.metric("Average Age", f"{avg_age} years")

        with col3:
            verification_rate = (members_df['is_verified'].sum() / len(members_df)) * 100 if len(members_df) > 0 else 0
            st.metric("Verification Rate", f"{verification_rate:.1f}%", f"{members_df['is_verified'].sum()} verified")

        with col4:
            unique_schools = members_df['school_name'].nunique()
            st.metric("Schools", unique_schools)

        with col5:
            avg_camps = round(members_df['youth_camps_attended'].mean(), 1) if len(members_df) > 0 else 0
            st.metric("Avg Camps", avg_camps)

        st.divider()

        st.subheader("📅 Registration Trends")
        col1, col2 = st.columns(2)

        with col1:
            daily_regs = members_df.set_index('submission_date').resample('D').size()
            if len(daily_regs) > 0:
                fig_daily = px.line(x=daily_regs.index, y=daily_regs.values,
                                    title="Daily Registration Trend",
                                    labels={'x': 'Date', 'y': 'Registrations'})
                fig_daily.update_layout(hovermode='x unified')
                st.plotly_chart(fig_daily, use_container_width=True)

        with col2:
            monthly_regs = members_df.set_index('submission_date').resample('ME').size()
            if len(monthly_regs) > 0:
                fig_monthly = px.bar(x=monthly_regs.index, y=monthly_regs.values,
                                     title="Monthly Registration Trend",
                                     labels={'x': 'Month', 'y': 'Registrations'},
                                     color=monthly_regs.values,
                                     color_continuous_scale='blues')
                st.plotly_chart(fig_monthly, use_container_width=True)

        st.subheader("👥 Demographics Analysis")
        col1, col2, col3 = st.columns(3)

        with col1:
            age_bins = [13, 16, 19, 22, 25, 31]
            age_labels = ['13-15', '16-18', '19-21', '22-24', '25-30']
            members_df['age_group'] = pd.cut(members_df['age'], bins=age_bins, labels=age_labels, right=False)
            age_dist = members_df['age_group'].value_counts().sort_index()
            if len(age_dist) > 0:
                fig_age = px.pie(values=age_dist.values, names=age_dist.index, title="Age Distribution",
                                color_discrete_sequence=px.colors.sequential.Blues_r)
                st.plotly_chart(fig_age, use_container_width=True)

        with col2:
            gender_dist = members_df['gender'].value_counts()
            if len(gender_dist) > 0:
                fig_gender = px.pie(values=gender_dist.values, names=gender_dist.index, title="Gender Distribution",
                                   color_discrete_sequence=['#667eea', '#764ba2'])
                st.plotly_chart(fig_gender, use_container_width=True)

        with col3:
            hall_dist = members_df['yoka_hall'].value_counts().head(10)
            if len(hall_dist) > 0:
                fig_hall = px.bar(x=hall_dist.index, y=hall_dist.values,
                                 title="YoKA Hall Distribution",
                                 color=hall_dist.values,
                                 color_continuous_scale='purples')
                st.plotly_chart(fig_hall, use_container_width=True)

        st.subheader("📊 Additional Insights")
        col_a, col_b = st.columns(2)

        with col_a:
            school_level_dist = members_df['school_level'].value_counts()
            if len(school_level_dist) > 0:
                fig_school = px.bar(x=school_level_dist.index, y=school_level_dist.values,
                                   title="Education Level Distribution",
                                   color=school_level_dist.values,
                                   color_continuous_scale='teal')
                st.plotly_chart(fig_school, use_container_width=True)

        with col_b:
            church_position_count = members_df['has_church_position'].sum() if 'has_church_position' in members_df.columns else 0
            fig_position = px.pie(values=[church_position_count, len(members_df) - church_position_count],
                                 names=['Has Church Position', 'No Church Position'],
                                 title="Church Leadership Participation",
                                 color_discrete_sequence=['#667eea', '#e0e0e0'])
            st.plotly_chart(fig_position, use_container_width=True)
    else:
        st.info("No data available for analytics. Start registering members to see insights!")

def branch_management_ui():
    if st.session_state.user_role not in ['super_admin', 'admin']:
        st.error("You don't have permission to manage branches")
        return

    st.header("🏢 Branch Management")

    tab1, tab2 = st.tabs(["All Branches", "Create New Branch"])

    with tab1:
        branches_df = get_all_branches(include_inactive=True)
        if not branches_df.empty:
            display_cols = ['id', 'branch_name', 'branch_code', 'location', 'contact_person', 'contact_phone', 'is_active']
            available_cols = [col for col in display_cols if col in branches_df.columns]
            st.dataframe(branches_df[available_cols], use_container_width=True)

            selected_branch = st.selectbox("Select Branch to Edit", branches_df['id'].tolist(),
                                          format_func=lambda x: branches_df[branches_df['id']==x]['branch_name'].iloc[0])
            if selected_branch:
                with st.expander("Edit Branch"):
                    branch_data = get_branch_by_id(selected_branch)
                    if branch_data:
                        new_branch_name = st.text_input("Branch Name", branch_data[1] if len(branch_data) > 1 else '')
                        new_branch_code = st.text_input("Branch Code", branch_data[2] if len(branch_data) > 2 else '')
                        new_location = st.text_input("Location", branch_data[3] if len(branch_data) > 3 else '')
                        new_contact_person = st.text_input("Contact Person", branch_data[4] if len(branch_data) > 4 else '')
                        new_contact_phone = st.text_input("Contact Phone", branch_data[5] if len(branch_data) > 5 else '')

                        col1, col2 = st.columns(2)
                        with col1:
                            if st.button("Update Branch"):
                                update_branch(selected_branch, new_branch_name, new_branch_code, new_location,
                                            new_contact_person, new_contact_phone, st.session_state.user_id)
                                st.success("Branch updated successfully!")
                                st.rerun()
                        with col2:
                            if st.button("Deactivate Branch", type="secondary"):
                                delete_branch(selected_branch, st.session_state.user_id)
                                st.warning("Branch deactivated!")
                                st.rerun()
        else:
            st.info("No branches found")

    with tab2:
        with st.form("create_branch_form"):
            col1, col2 = st.columns(2)
            with col1:
                branch_name = st.text_input("Branch Name *")
                branch_code = st.text_input("Branch Code *")
                location = st.text_input("Location *")
            with col2:
                contact_person = st.text_input("Contact Person *")
                contact_phone = st.text_input("Contact Phone *", help="Format: 024XXXXXXX")

            if st.form_submit_button("Create Branch"):
                if all([branch_name, branch_code, location, contact_person, contact_phone]):
                    if not validate_ghana_phone(contact_phone):
                        st.error("Invalid Ghana phone number format")
                    else:
                        success, result = create_branch(branch_name, branch_code, location, contact_person,
                                                       contact_phone, st.session_state.user_id)
                        if success:
                            st.success(f"Branch {branch_name} created successfully!")
                            st.rerun()
                        else:
                            st.error(result)
                else:
                    st.error("Please fill all required fields")

def audit_log_ui():
    if st.session_state.user_role != 'super_admin':
        st.error("Only Super Admin can access audit logs")
        return

    st.title("📜 Audit Log")

    conn = sqlite3.connect('kumasi_yoka_registration.db')
    df = pd.read_sql_query("SELECT * FROM audit_log ORDER BY timestamp DESC LIMIT 1000", conn)
    conn.close()

    if not df.empty:
        display_cols = ['timestamp', 'username', 'action', 'entity_type', 'entity_id', 'details']
        available_cols = [col for col in display_cols if col in df.columns]
        st.dataframe(df[available_cols], use_container_width=True)

        csv = df.to_csv(index=False)
        st.download_button("Download Audit Log", csv,
                         f"audit_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
    else:
        st.info("No audit logs found")

def registration_form_ui():
    st.title("⛪ YoKA Registration Form")

    # Animated success toast CSS
    st.markdown("""
    <style>
    @keyframes slideInDown {
        from { opacity: 0; transform: translateY(-60px) scale(0.95); }
        to   { opacity: 1; transform: translateY(0)    scale(1);    }
    }
    @keyframes fadeOutUp {
        from { opacity: 1; transform: translateY(0)    scale(1);    }
        to   { opacity: 0; transform: translateY(-60px) scale(0.95); }
    }
    .success-toast {
        position: fixed;
        top: 20px;
        left: 50%;
        transform: translateX(-50%);
        z-index: 9999;
        background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
        color: white;
        padding: 1.1rem 2.2rem;
        border-radius: 16px;
        box-shadow: 0 8px 32px rgba(17,153,142,0.35);
        font-size: 1.15rem;
        font-weight: 700;
        display: flex;
        align-items: center;
        gap: 0.7rem;
        animation: slideInDown 0.5s ease forwards,
                   fadeOutUp   0.6s ease 2.4s forwards;
        letter-spacing: 0.01em;
        white-space: nowrap;
    }
    .success-toast .toast-icon {
        font-size: 1.6rem;
        animation: bounceIn 0.6s ease 0.3s both;
    }
    @keyframes bounceIn {
        0%   { transform: scale(0);   }
        60%  { transform: scale(1.3); }
        100% { transform: scale(1);   }
    }
    </style>
    """, unsafe_allow_html=True)

    if st.session_state.get('registration_success'):
        st.markdown("""
        <div class="success-toast">
            <span class="toast-icon">✅</span>
            Member registered successfully! Form is ready for next entry.
        </div>
        """, unsafe_allow_html=True)

    # ── KEY FIX: use the counter as the form key so the form resets after submit ──
    with st.form(f"registration_form_{st.session_state.form_key_counter}"):
        st.header("📋 Personal Information")
        col1, col2 = st.columns(2)

        with col1:
            official_name = st.text_input("Full Name *")
            gender = st.selectbox("Gender *", ["Male", "Female", "Other"])
            date_of_birth = st.date_input("Date of Birth",
                                         min_value=date(1950, 1, 1),
                                         max_value=date.today(),
                                         value=date(2005, 1, 1))
            residence = st.text_input("Residential Address *")
            email = st.text_input("Email Address")
            profile_picture = st.file_uploader("Profile Picture", type=['jpg', 'jpeg', 'png'])

        with col2:
            if date_of_birth:
                try:
                    calculated_age = calculate_age_from_dob(date_of_birth)
                    if calculated_age < 13:
                        st.warning(f"⚠️ Age ({calculated_age}) is below the allowed range (13-30 years).")
                    elif calculated_age > 30:
                        st.warning(f"⚠️ Age ({calculated_age}) is above the allowed range (13-30 years).")
                    else:
                        st.success(f"✅ Age: {calculated_age} years")
                except:
                    pass

            active_phone = st.text_input("Active Phone Number(s) *", help="Format: 024XXXXXXX or 054XXXXXXX")
            emergency_contact_name = st.text_input("Emergency Contact Name")
            emergency_contact_phone = st.text_input("Emergency Contact Phone", help="Format: 024XXXXXXX")
            medical_conditions = st.text_area("Medical Conditions/Allergies (if any)")

        st.header("🎓 School Information")

        school_level = st.selectbox("Education Level *", ["SHS", "Tertiary", "Graduate", "Other"])

        school_name = ""
        school_class = ""
        school_house = ""
        residence_status = "Day Student"
        residence_name = ""
        programme = ""
        courses = ""
        hall_or_hostel = ""
        form = ""

        if school_level == "Tertiary":
            st.subheader("Tertiary Education Details")
            col_t1, col_t2 = st.columns(2)
            with col_t1:
                school_name = st.text_input("University/College Name *")
                programme = st.text_input("Programme of Study *")
                level = st.selectbox("Current Level/Year *", ["Level 100", "Level 200", "Level 300", "Level 400", "Postgraduate"])
            with col_t2:
                courses = st.text_area("Main Courses/Subjects", help="List your major courses")
                hall_or_hostel = st.text_input("Hall/Hostel Name")
                residence_status = st.selectbox("Residence Status *", ["Day Student", "Hall Resident", "Hostel Resident", "Off-Campus"])
            school_class = level
            school_house = hall_or_hostel
            residence_name = st.text_input("Residence/Hall Name") if residence_status != "Day Student" else ""

        elif school_level == "SHS":
            st.subheader("Secondary School Details")
            col_s1, col_s2 = st.columns(2)
            with col_s1:
                school_name = st.text_input("School Name *")
                form = st.selectbox("Form/Year *", ["Form 1", "Form 2", "Form 3"])
                programme = st.selectbox("Programme *", ["General Arts", "General Science", "Business", "Visual Arts", "Home Economics", "Technical"])
            with col_s2:
                courses = st.text_area("Subjects Offered", help="List your main subjects")
                residence_status = st.selectbox("Residence Status *", ["Day Student", "Boarder"])
            school_class = form
            school_house = ""
            residence_name = st.text_input("Hostel Name") if residence_status == "Boarder" else ""
        else:
            school_name = st.text_input("School Name *")
            school_class = st.text_input("Current Class/Level *")
            programme = st.text_input("Programme (if applicable)")
            courses = st.text_area("Courses/Subjects")
            residence_status = st.selectbox("Residence Status *", ["Day Student", "Boarder", "Off-Campus"])
            residence_name = st.text_input("Residence/Hostel Name") if residence_status != "Day Student" else ""
            school_house = st.text_input("House (if applicable)")

        st.header("🎭 Talents & Interests")
        col5, col6 = st.columns(2)
        with col5:
            talents = st.multiselect("Talents/Skills",
                                    ["Singing", "Dancing", "Drama", "Instrumental", "Public Speaking",
                                     "Sports", "Teaching", "Technical", "Writing", "Graphic Design", "Other"])
        with col6:
            interests = st.multiselect("Areas of Interest",
                                      ["Evangelism", "Teaching", "Technical Support", "Event Planning",
                                       "Counseling", "Media", "Worship Team", "Ushering", "Protocol", "Other"])

        st.header("⛪ Church Information")

        if st.session_state.user_role == 'branch_executive':
            church_branch = st.text_input("Church Branch", value=st.session_state.branch_name or "", disabled=True)
        else:
            branches_df = get_all_branches()
            if not branches_df.empty:
                branch_options = branches_df['branch_name'].tolist()
                church_branch = st.selectbox("Church Branch *", branch_options)
            else:
                church_branch = st.text_input("Church Branch *")

        yoka_hall = st.selectbox("YoKA Hall *", ["Hall A", "Hall B", "Hall C", "Hall D", "Hall E", "Other"])
        if yoka_hall == "Other":
            yoka_hall = st.text_input("Specify Hall")

        youth_camps = st.slider("Number of YoKA Youth Camps Attended *", min_value=0, max_value=20, value=0)

        st.subheader("🙏 Church Leadership & Service")
        st.caption("Please indicate if you currently hold or have formerly held a church position")

        has_church_position = st.radio("Do you currently hold or have you formerly held a church position?",
                                       ["No", "Yes - Presently Holding", "Yes - Formerly Held"],
                                       index=0)

        church_position_status = ""
        church_position_type = ""
        church_position_name = ""
        church_position_duration = ""

        if has_church_position != "No":
            col_pos1, col_pos2 = st.columns(2)
            with col_pos1:
                church_position_type = st.selectbox("Position Type", ["Seven Member", "YoKA Executive", "Other"])
                if church_position_type == "Other":
                    church_position_name = st.text_input("Please specify the position name")
                else:
                    church_position_name = church_position_type
            with col_pos2:
                church_position_duration = st.text_input("How long have you been/had been in this position?",
                                                        placeholder="e.g., 2 years, 6 months, 1 year 3 months")

            if has_church_position == "Yes - Presently Holding":
                church_position_status = "Present"
            else:
                church_position_status = "Former"

        st.header("💼 Work Information")
        col_work1, col_work2 = st.columns(2)
        with col_work1:
            work_status = st.selectbox("Work Status", ["None", "Apprenticeship", "Internship", "Permanent Staff", "Part-time"])
        with col_work2:
            work_type = ""
            if work_status != "None":
                work_type = st.selectbox("Type of Work", ["Formal Employment", "Self-Employed", "Freelance", "Volunteer", "Other"])

        work_name = work_position = work_location = ""
        work_experience_years = 0

        if work_status != "None":
            col_work3, col_work4 = st.columns(2)
            with col_work3:
                work_name = st.text_input("Name of Workplace/Organization")
                work_position = st.text_input("Your Position/Title")
            with col_work4:
                work_location = st.text_input("Work Location (City/Country)")
                work_experience_years = st.number_input("Number of Years of Experience", min_value=0, max_value=30, value=0)

        st.header("🌍 Diaspora Information")
        is_diaspora = st.checkbox("Are you currently in the diaspora (outside Ghana)?")

        diaspora_country = diaspora_status = diaspora_job = diaspora_school = diaspora_education_level = ""

        if is_diaspora:
            col_dias1, col_dias2 = st.columns(2)
            with col_dias1:
                diaspora_country = st.text_input("Which country are you in? *")
                diaspora_status = st.selectbox("Status in Diaspora", ["Working", "Studying", "Both"])

            with col_dias2:
                if diaspora_status in ["Working", "Both"]:
                    diaspora_job = st.text_input("Current Job/Profession")

                if diaspora_status in ["Studying", "Both"]:
                    diaspora_school = st.text_input("School/Institution")
                    diaspora_education_level = st.selectbox("Level of Education",
                                                           ["Undergraduate", "Graduate", "Masters", "PhD", "Other"])

        st.header("👨‍👩‍👧 Parent/Guardian Information")
        st.info("Please provide information for at least one parent or guardian")

        st.subheader("👩 Mother's Information")
        col_m1, col_m2 = st.columns(2)
        with col_m1:
            mother_name = st.text_input("Mother's Full Name")
            mother_phone = st.text_input("Mother's Phone Number", help="Format: 024XXXXXXX")
        with col_m2:
            mother_occupation = st.text_input("Mother's Occupation")

        st.divider()

        st.subheader("👨 Father's Information")
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            father_name = st.text_input("Father's Full Name")
            father_phone = st.text_input("Father's Phone Number", help="Format: 024XXXXXXX")
        with col_f2:
            father_occupation = st.text_input("Father's Occupation")

        st.divider()

        st.subheader("👤 Guardian Information (if applicable)")
        st.caption("Provide guardian information if the member is not living with parents or if parents are unavailable")

        col_g1, col_g2, col_g3 = st.columns(3)
        with col_g1:
            guardian_name = st.text_input("Guardian's Full Name")
        with col_g2:
            guardian_phone = st.text_input("Guardian's Phone Number", help="Format: 024XXXXXXX")
        with col_g3:
            guardian_relationship = st.text_input("Relationship to Member")
        guardian_occupation = st.text_input("Guardian's Occupation")

        submitted = st.form_submit_button("Register Member", use_container_width=True, type="primary")

        if submitted:
            age_calculated = calculate_age_from_dob(date_of_birth) if date_of_birth else 18

            if age_calculated < 13:
                st.error(f"Member must be at least 13 years old. Current age: {age_calculated}")
            elif age_calculated > 30:
                st.error(f"Member must be at most 30 years old. Current age: {age_calculated}")
            else:
                phone_errors = []
                if active_phone and not validate_ghana_phone(active_phone):
                    phone_errors.append("Member's phone number")
                if emergency_contact_phone and not validate_ghana_phone(emergency_contact_phone):
                    phone_errors.append("Emergency contact phone")
                if mother_phone and not validate_ghana_phone(mother_phone):
                    phone_errors.append("Mother's phone")
                if father_phone and not validate_ghana_phone(father_phone):
                    phone_errors.append("Father's phone")
                if guardian_phone and not validate_ghana_phone(guardian_phone):
                    phone_errors.append("Guardian's phone")

                if phone_errors:
                    st.error(f"Invalid Ghana phone number format for: {', '.join(phone_errors)}")
                elif email and not validate_email(email):
                    st.error("Invalid email format")
                else:
                    has_parent_guardian = (mother_name and mother_phone) or (father_name and father_phone) or (guardian_name and guardian_phone)

                    if all([official_name, residence, active_phone, school_name, residence_status, yoka_hall, gender, church_branch]) and has_parent_guardian:
                        talents_str = ', '.join(talents) if talents else ''
                        interests_str = ', '.join(interests) if interests else ''

                        data = {
                            'official_name': official_name,
                            'date_of_birth': date_of_birth.strftime("%Y-%m-%d") if date_of_birth else "",
                            'age': age_calculated,
                            'residence': residence,
                            'active_phone': active_phone,
                            'email': email if email else '',
                            'school_name': school_name,
                            'school_level': school_level,
                            'school_class': school_class if school_class else "",
                            'school_house': school_house if school_house else "",
                            'residence_status': residence_status,
                            'residence_name': residence_name if residence_status in ["Boarder", "Hall Resident", "Hostel Resident"] else "",
                            'church_branch': church_branch,
                            'yoka_hall': yoka_hall,
                            'youth_camps_attended': youth_camps,
                            'has_church_position': has_church_position != "No",
                            'church_position_status': church_position_status,
                            'church_position_type': church_position_type,
                            'church_position_name': church_position_name,
                            'church_position_duration': church_position_duration,
                            'work_status': work_status if work_status != "None" else "",
                            'work_type': work_type,
                            'work_name': work_name,
                            'work_position': work_position,
                            'work_location': work_location,
                            'work_experience_years': work_experience_years,
                            'is_diaspora': is_diaspora,
                            'diaspora_country': diaspora_country,
                            'diaspora_status': diaspora_status,
                            'diaspora_job': diaspora_job,
                            'diaspora_school': diaspora_school,
                            'diaspora_education_level': diaspora_education_level,
                            'mother_name': mother_name,
                            'mother_phone': mother_phone,
                            'mother_occupation': mother_occupation,
                            'father_name': father_name,
                            'father_phone': father_phone,
                            'father_occupation': father_occupation,
                            'guardian_name': guardian_name,
                            'guardian_phone': guardian_phone,
                            'guardian_relationship': guardian_relationship,
                            'guardian_occupation': guardian_occupation,
                            'gender': gender,
                            'emergency_contact_name': emergency_contact_name,
                            'emergency_contact_phone': emergency_contact_phone,
                            'medical_conditions': medical_conditions,
                            'talents': talents_str,
                            'interests': interests_str,
                            'programme': programme if programme else "",
                            'courses': courses if courses else "",
                            'hall_or_hostel': hall_or_hostel if school_level == "Tertiary" else "",
                            'form': form if school_level == "SHS" else ""
                        }

                        try:
                            save_member(data, profile_picture, st.session_state.user_id)
                            # ── KEY FIX: increment counter so form renders fresh on rerun ──
                            st.session_state.form_key_counter += 1
                            st.session_state.registration_success = True
                            st.balloons()
                            time.sleep(3)
                            st.session_state.registration_success = False
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error saving member: {str(e)}")
                    else:
                        missing_fields = []
                        if not official_name: missing_fields.append("Full Name")
                        if not residence: missing_fields.append("Residential Address")
                        if not active_phone: missing_fields.append("Phone Number")
                        if not school_name: missing_fields.append("School Name")
                        if not residence_status: missing_fields.append("Residence Status")
                        if not church_branch: missing_fields.append("Church Branch")
                        if not yoka_hall: missing_fields.append("YoKA Hall")
                        if not gender: missing_fields.append("Gender")
                        if not has_parent_guardian: missing_fields.append("At least one Parent/Guardian")

                        if missing_fields:
                            st.error(f"Please fill in the following required fields: {', '.join(missing_fields)}")

def view_members_ui():
    st.title("📋 Registered Members")

    if st.session_state.user_role == 'branch_executive':
        members_df = get_members_by_role(st.session_state.user_role, st.session_state.user_branch_id)
        st.info(f"Showing members from {st.session_state.branch_name} branch")
    else:
        members_df = get_members_by_role(st.session_state.user_role)
        branches_df = get_all_branches()
        if not branches_df.empty:
            branch_options = ["All Branches"] + branches_df['branch_name'].tolist()
            selected_branch = st.selectbox("Filter by Branch", branch_options)
            if selected_branch != "All Branches":
                members_df = members_df[members_df['church_branch'] == selected_branch]

    if not members_df.empty:
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            st.metric("Total Members", len(members_df))
        with col2:
            st.metric("Average Age", round(members_df['age'].mean(), 1) if len(members_df) > 0 else 0)
        with col3:
            st.metric("Youth Camps (Avg)", round(members_df['youth_camps_attended'].mean(), 1) if len(members_df) > 0 else 0)
        with col4:
            verified_count = members_df['is_verified'].sum() if 'is_verified' in members_df.columns else 0
            st.metric("Verified", verified_count)
        with col5:
            photo_count = members_df['profile_picture'].notna().sum() if 'profile_picture' in members_df.columns else 0
            st.metric("With Photos", photo_count)

        search_term = st.text_input("🔍 Search by name or phone")
        if search_term:
            members_df = members_df[members_df['official_name'].str.contains(search_term, case=False, na=False) |
                                   members_df['active_phone'].str.contains(search_term, case=False, na=False)]

        display_cols = ['id', 'official_name', 'age', 'gender', 'active_phone', 'church_branch', 'yoka_hall', 'submission_date', 'is_verified']
        available_cols = [col for col in display_cols if col in members_df.columns]

        with st.expander("Bulk Operations"):
            st.write("Select multiple members for bulk actions")
            selected_ids = st.multiselect("Select members", members_df['id'].tolist(),
                                          format_func=lambda x: f"{x} - {members_df[members_df['id']==x]['official_name'].iloc[0]}")
            if selected_ids:
                col_b1, col_b2 = st.columns(2)
                with col_b1:
                    if st.button("🗑️ Delete Selected", type="secondary"):
                        if st.session_state.user_role == 'super_admin':
                            count = delete_bulk_members(selected_ids, st.session_state.user_id)
                            st.success(f"Deleted {count} members successfully!")
                            st.rerun()
                        else:
                            st.error("Only Super Admin can perform bulk delete")
                with col_b2:
                    if st.button("✅ Verify Selected"):
                        if st.session_state.user_role in ['super_admin', 'admin']:
                            for member_id in selected_ids:
                                verify_member(member_id, st.session_state.user_id)
                            st.success(f"Verified {len(selected_ids)} members!")
                            st.rerun()
                        else:
                            st.error("Only Admin and Super Admin can verify members")

        st.dataframe(members_df[available_cols], use_container_width=True)

        if len(members_df) > 0:
            selected_id = st.selectbox("Select Member to View/Edit Details", members_df['id'].tolist())
            if selected_id:
                member = get_member_by_id(selected_id, st.session_state.user_role, st.session_state.user_branch_id)
                if member:
                    tab1, tab2 = st.tabs(["View Details", "Edit Member"])

                    with tab1:
                        col1, col2 = st.columns([1, 2])
                        with col1:
                            if len(member) > 7 and member[7]:
                                try:
                                    st.image(member[7], caption="Profile Picture", use_container_width=True)
                                except:
                                    st.info("No profile picture available")
                            else:
                                st.info("No profile picture uploaded")

                        with col2:
                            st.write(f"**Name:** {member[1] if len(member) > 1 else 'N/A'}")
                            st.write(f"**Gender:** {member[52] if len(member) > 52 else 'N/A'}")
                            st.write(f"**Age:** {member[3] if len(member) > 3 else 'N/A'}")
                            st.write(f"**Phone:** {member[5] if len(member) > 5 else 'N/A'}")
                            st.write(f"**Email:** {member[6] if len(member) > 6 else 'N/A'}")
                            st.write(f"**Residence:** {member[4] if len(member) > 4 else 'N/A'}")
                            st.write(f"**Branch:** {member[14] if len(member) > 14 else 'N/A'}")
                            st.write(f"**Hall:** {member[16] if len(member) > 16 else 'N/A'}")
                            st.write(f"**School Level:** {member[9] if len(member) > 9 else 'N/A'}")
                            st.write(f"**School:** {member[8] if len(member) > 8 else 'N/A'}")
                            st.write(f"**Class/Form:** {member[10] if len(member) > 10 else 'N/A'}")
                            st.write(f"**Registered:** {member[45] if len(member) > 45 else 'N/A'}")

                        if len(member) > 18 and member[18]:
                            st.subheader("⛪ Church Position")
                            position_status = member[19] if len(member) > 19 else 'N/A'
                            status_icon = "✅" if position_status == "Present" else "📜"
                            st.write(f"{status_icon} **Status:** {position_status} Position")
                            st.write(f"**Position:** {member[21] if len(member) > 21 else 'N/A'}")
                            st.write(f"**Duration:** {member[22] if len(member) > 22 else 'N/A'}")

                        if len(member) > 49 and member[49]:
                            st.success("✅ Verified Member")
                        else:
                            st.warning("⏳ Pending Verification")
                            if st.session_state.user_role in ['super_admin', 'admin']:
                                if st.button("Verify Member"):
                                    verify_member(selected_id, st.session_state.user_id)
                                    st.success("Member verified!")
                                    st.rerun()

                    with tab2:
                        st.subheader("✏️ Edit Member Information")
                        with st.form("edit_member_form"):
                            col1, col2 = st.columns(2)
                            with col1:
                                edit_name = st.text_input("Full Name", member[1] if len(member) > 1 else "")
                                edit_gender = st.selectbox("Gender", ["Male", "Female", "Other"],
                                                          index=["Male", "Female", "Other"].index(member[52]) if len(member) > 52 and member[52] in ["Male", "Female", "Other"] else 0)
                                edit_phone = st.text_input("Phone Number", member[5] if len(member) > 5 else "")
                                edit_residence = st.text_input("Residence", member[4] if len(member) > 4 else "")
                                edit_email = st.text_input("Email", member[6] if len(member) > 6 else "")
                            with col2:
                                edit_school = st.text_input("School", member[8] if len(member) > 8 else "")
                                edit_hall = st.text_input("YoKA Hall", member[16] if len(member) > 16 else "")
                                edit_camps = st.number_input("Youth Camps Attended",
                                                            value=int(member[17]) if len(member) > 17 and member[17] else 0,
                                                            min_value=0, max_value=20)
                                new_profile_picture = st.file_uploader("Update Profile Picture", type=['jpg', 'jpeg', 'png'])

                            col3, col4 = st.columns(2)
                            with col3:
                                edit_mother_name = st.text_input("Mother's Name", member[35] if len(member) > 35 else "")
                                edit_mother_phone = st.text_input("Mother's Phone", member[36] if len(member) > 36 else "")
                            with col4:
                                edit_father_name = st.text_input("Father's Name", member[38] if len(member) > 38 else "")
                                edit_father_phone = st.text_input("Father's Phone", member[39] if len(member) > 39 else "")

                            if st.form_submit_button("💾 Save Changes"):
                                if edit_phone and not validate_ghana_phone(edit_phone):
                                    st.error("Invalid Ghana phone number format")
                                elif edit_email and not validate_email(edit_email):
                                    st.error("Invalid email format")
                                else:
                                    edit_data = {
                                        'official_name': edit_name,
                                        'date_of_birth': member[2] if len(member) > 2 else "",
                                        'age': member[3] if len(member) > 3 else 18,
                                        'residence': edit_residence,
                                        'active_phone': edit_phone,
                                        'email': edit_email,
                                        'school_name': edit_school,
                                        'school_level': member[9] if len(member) > 9 else "SHS",
                                        'school_class': member[10] if len(member) > 10 else "",
                                        'school_house': member[11] if len(member) > 11 else "",
                                        'residence_status': member[12] if len(member) > 12 else "Day Student",
                                        'residence_name': member[13] if len(member) > 13 else "",
                                        'church_branch': member[14] if len(member) > 14 else "",
                                        'yoka_hall': edit_hall,
                                        'youth_camps_attended': edit_camps,
                                        'has_church_position': bool(member[18]) if len(member) > 18 else False,
                                        'church_position_status': member[19] if len(member) > 19 else "",
                                        'church_position_type': member[20] if len(member) > 20 else "",
                                        'church_position_name': member[21] if len(member) > 21 else "",
                                        'church_position_duration': member[22] if len(member) > 22 else "",
                                        'work_status': member[23] if len(member) > 23 else "",
                                        'work_type': member[24] if len(member) > 24 else "",
                                        'work_name': member[25] if len(member) > 25 else "",
                                        'work_position': member[26] if len(member) > 26 else "",
                                        'work_location': member[27] if len(member) > 27 else "",
                                        'work_experience_years': member[28] if len(member) > 28 else 0,
                                        'is_diaspora': member[29] if len(member) > 29 else 0,
                                        'diaspora_country': member[30] if len(member) > 30 else "",
                                        'diaspora_status': member[31] if len(member) > 31 else "",
                                        'diaspora_job': member[32] if len(member) > 32 else "",
                                        'diaspora_school': member[33] if len(member) > 33 else "",
                                        'diaspora_education_level': member[34] if len(member) > 34 else "",
                                        'mother_name': edit_mother_name,
                                        'mother_phone': edit_mother_phone,
                                        'mother_occupation': member[37] if len(member) > 37 else "",
                                        'father_name': edit_father_name,
                                        'father_phone': edit_father_phone,
                                        'father_occupation': member[40] if len(member) > 40 else "",
                                        'guardian_name': member[41] if len(member) > 41 else "",
                                        'guardian_phone': member[42] if len(member) > 42 else "",
                                        'guardian_relationship': member[43] if len(member) > 43 else "",
                                        'guardian_occupation': member[44] if len(member) > 44 else "",
                                        'gender': edit_gender,
                                        'emergency_contact_name': member[53] if len(member) > 53 else "",
                                        'emergency_contact_phone': member[54] if len(member) > 54 else "",
                                        'medical_conditions': member[55] if len(member) > 55 else "",
                                        'talents': member[56] if len(member) > 56 else "",
                                        'interests': member[57] if len(member) > 57 else "",
                                        'programme': member[58] if len(member) > 58 else "",
                                        'courses': member[59] if len(member) > 59 else "",
                                        'hall_or_hostel': member[60] if len(member) > 60 else "",
                                        'form': member[61] if len(member) > 61 else ""
                                    }
                                    try:
                                        update_member(selected_id, edit_data, st.session_state.user_id, new_profile_picture)
                                        st.success("Member updated successfully!")
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"Error updating member: {str(e)}")

                        if st.session_state.user_role == 'super_admin':
                            if st.button("🗑️ Delete Member", type="secondary"):
                                success, message = delete_member(selected_id, st.session_state.user_id,
                                                                st.session_state.user_role, st.session_state.user_branch_id)
                                if success:
                                    st.success(message)
                                    st.rerun()
                                else:
                                    st.error(message)
    else:
        st.info("No members registered yet")

def user_management_ui():
    if st.session_state.user_role != 'super_admin':
        st.error("Only Super Admin can access user management")
        return

    st.header("👥 User Management")

    tab1, tab2 = st.tabs(["All Users", "Create New User"])

    with tab1:
        users_df = get_all_users()
        if not users_df.empty:
            display_cols = ['id', 'username', 'full_name', 'email', 'phone', 'role', 'assigned_branch', 'is_active']
            available_cols = [col for col in display_cols if col in users_df.columns]
            st.dataframe(users_df[available_cols], use_container_width=True)

            selected_user = st.selectbox("Select User to Edit", users_df['id'].tolist(),
                                        format_func=lambda x: f"{x} - {users_df[users_df['id']==x]['full_name'].iloc[0]}")
            if selected_user:
                with st.expander("Edit/Delete User"):
                    user_data = get_user_by_id(selected_user)
                    if user_data:
                        new_full_name = st.text_input("Full Name", user_data[3] if len(user_data) > 3 else '')
                        new_email = st.text_input("Email", user_data[4] if len(user_data) > 4 else '')
                        new_phone = st.text_input("Phone", user_data[5] if len(user_data) > 5 else '')

                        role_options = ['super_admin', 'admin', 'branch_executive']
                        current_role_index = role_options.index(user_data[6]) if user_data[6] in role_options else 0
                        new_role = st.selectbox("Role", role_options, index=current_role_index)

                        branches_df = get_all_branches()
                        if not branches_df.empty:
                            branch_options = {row['id']: row['branch_name'] for _, row in branches_df.iterrows()}
                            current_branch = user_data[7] if len(user_data) > 7 else None
                            branch_keys = list(branch_options.keys())
                            branch_index = branch_keys.index(current_branch) if current_branch in branch_keys else 0
                            new_branch = st.selectbox("Assigned Branch", branch_keys,
                                                     format_func=lambda x: branch_options.get(x, 'None'),
                                                     index=branch_index)
                        else:
                            new_branch = None
                            st.warning("No branches available")

                        is_active = st.checkbox("Active", user_data[8] == 1 if len(user_data) > 8 else True)

                        col1, col2, col3 = st.columns(3)
                        with col1:
                            if st.button("Update User"):
                                if new_branch:
                                    if new_phone and not validate_ghana_phone(new_phone):
                                        st.error("Invalid Ghana phone number format")
                                    else:
                                        success, message = update_user(selected_user, new_full_name, new_email, new_phone,
                                                                      new_role, new_branch, is_active, st.session_state.user_id)
                                        if success:
                                            st.success(message)
                                            st.rerun()
                                        else:
                                            st.error(message)

                        with col2:
                            new_password = st.text_input("New Password (leave blank to keep current)", type="password")
                            if st.button("Reset Password") and new_password:
                                reset_user_password(selected_user, new_password, st.session_state.user_id)
                                st.success("Password reset and email sent successfully!")

                        with col3:
                            if st.button("Delete User", type="secondary"):
                                if selected_user == st.session_state.user_id:
                                    st.error("Cannot delete your own account")
                                else:
                                    delete_user(selected_user, st.session_state.user_id)
                                    st.success("User deleted successfully!")
                                    st.rerun()
        else:
            st.info("No users found")

    with tab2:
        with st.form("create_user_form"):
            col1, col2 = st.columns(2)
            with col1:
                username = st.text_input("Username *")
                full_name = st.text_input("Full Name *")
                email = st.text_input("Email *")
                phone = st.text_input("Phone *", help="Format: 024XXXXXXX")
            with col2:
                password = st.text_input("Password *", type="password")
                role = st.selectbox("Role *", ['super_admin', 'admin', 'branch_executive'])

                branches_df = get_all_branches()
                if not branches_df.empty:
                    branch_options = {row['id']: row['branch_name'] for _, row in branches_df.iterrows()}
                    assigned_branch = st.selectbox("Assigned Branch *", list(branch_options.keys()),
                                                  format_func=lambda x: branch_options[x])
                else:
                    assigned_branch = None
                    st.error("No branches available. Please create a branch first.")

            if st.form_submit_button("Create User"):
                if all([username, full_name, email, phone, password, assigned_branch]):
                    if not validate_ghana_phone(phone):
                        st.error("Invalid Ghana phone number format")
                    elif email and not validate_email(email):
                        st.error("Invalid email format")
                    else:
                        success, result = create_user(username, password, full_name, email, phone, role,
                                                     assigned_branch, st.session_state.user_id)
                        if success:
                            st.success(f"User {username} created successfully! Credentials sent to their email.")
                            st.rerun()
                        else:
                            st.error(result)
                else:
                    st.error("Please fill all required fields")

# ==================== MAIN APPLICATION ====================
def init_session_state():
    if 'logged_in' not in st.session_state:
        st.session_state.logged_in = False
    if 'user_id' not in st.session_state:
        st.session_state.user_id = None
    if 'username' not in st.session_state:
        st.session_state.username = None
    if 'user_role' not in st.session_state:
        st.session_state.user_role = None
    if 'user_branch_id' not in st.session_state:
        st.session_state.user_branch_id = None
    if 'branch_name' not in st.session_state:
        st.session_state.branch_name = None
    if 'selected_page' not in st.session_state:
        st.session_state.selected_page = "📝 Registration Form"
    if 'registration_success' not in st.session_state:
        st.session_state.registration_success = False
    # ── KEY FIX: counter used as the registration form's key to force a full reset ──
    if 'form_key_counter' not in st.session_state:
        st.session_state.form_key_counter = 0

def login_page():
    settings_df = get_system_settings()
    settings = {row['setting_key']: row['setting_value'] for _, row in settings_df.iterrows()} if not settings_df.empty else {}
    system_name = settings.get('system_name', 'Kumasi District YoKA Registration System')
    favicon = settings.get('favicon', '⛪')

    st.markdown("""
    <style>
        .stApp {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        }
        .login-container {
            max-width: 450px;
            margin: 5rem auto;
            padding: 2rem;
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            animation: fadeInUp 0.6s ease-out;
        }
        @keyframes fadeInUp {
            from { opacity: 0; transform: translateY(30px); }
            to { opacity: 1; transform: translateY(0); }
        }
        .login-header {
            text-align: center;
            margin-bottom: 2rem;
        }
        .login-header h1 {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            margin-bottom: 0.5rem;
        }
    </style>
    """, unsafe_allow_html=True)

    st.markdown(f"""
    <div class="login-container">
        <div class="login-header">
            <div style="font-size: 4rem;">{favicon}</div>
            <h1>{system_name}</h1>
            <p>Youth of Kristo Asafo Registration System</p>
            <p style="color: #666; font-size: 0.9rem;">Kumasi District</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        with st.form("login_form"):
            username = st.text_input("Username", placeholder="Enter your username")
            password = st.text_input("Password", type="password", placeholder="Enter your password")
            submit = st.form_submit_button("🔐 Login", use_container_width=True)

            if submit:
                user_info = check_login(username, password)
                if user_info:
                    st.session_state.logged_in = True
                    st.session_state.user_id = user_info['id']
                    st.session_state.username = user_info['username']
                    st.session_state.user_role = user_info['role']
                    st.session_state.user_branch_id = user_info.get('assigned_branch_id')
                    st.session_state.branch_name = user_info.get('branch_name')

                    st.success(f"Welcome back, {user_info['full_name']}! 🎉")
                    st.rerun()
                else:
                    st.error("❌ Invalid username or password. Please try again.")

def main_dashboard():
    settings_df = get_system_settings()
    settings = {row['setting_key']: row['setting_value'] for _, row in settings_df.iterrows()} if not settings_df.empty else {}
    system_name = settings.get('system_name', 'Kumasi District YoKA Registration System')

    st.sidebar.markdown("""
    <div style="text-align: center; padding: 1rem 0; border-bottom: 2px solid rgba(255,255,255,0.2); margin-bottom: 1rem;">
        <div style="font-size: 3rem;">⛪</div>
        <h2 style="margin: 0.5rem 0 0 0; font-size: 1.2rem; color: white;">YoKA Registration</h2>
        <p style="margin: 0; font-size: 0.8rem; opacity: 0.9; color: rgba(255,255,255,0.9);">Kumasi District</p>
    </div>
    """, unsafe_allow_html=True)

    st.sidebar.markdown(f"""
    <div style="background: rgba(255,255,255,0.15); padding: 1rem; border-radius: 12px; margin-bottom: 1rem; backdrop-filter: blur(5px);">
        <div style="display: flex; align-items: center; gap: 10px;">
            <div style="font-size: 2rem;">👤</div>
            <div>
                <div style="font-weight: 600; font-size: 0.9rem; color: white;">{st.session_state.username}</div>
                <div style="font-size: 0.75rem; opacity: 0.9; color: rgba(255,255,255,0.9);">{st.session_state.user_role.replace('_', ' ').title()}</div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.sidebar.markdown('<p style="color: white; font-weight: 600; margin-bottom: 0.5rem;">📋 Navigation</p>', unsafe_allow_html=True)

    nav_options = ["📝 Registration Form", "👥 View Members", "📊 Analytics Dashboard", "📥 Export Data"]

    if st.session_state.user_role in ['super_admin', 'admin']:
        nav_options.extend(["📤 Import Data", "🏢 Branch Management"])

    if st.session_state.user_role == 'super_admin':
        nav_options.extend(["👥 User Management", "📧 Email Settings", "🎨 Theme Customizer", "📜 Audit Log"])

    for option in nav_options:
        if st.sidebar.button(option, key=f"nav_{option}", use_container_width=True):
            st.session_state.selected_page = option
            st.rerun()

    st.sidebar.markdown("---")
    st.sidebar.markdown('<p style="color: white; font-weight: 600; margin-bottom: 0.5rem;">📊 System Stats</p>', unsafe_allow_html=True)

    conn = sqlite3.connect('kumasi_yoka_registration.db')
    members_count = pd.read_sql_query("SELECT COUNT(*) as count FROM members", conn).iloc[0]['count']
    branches_count = pd.read_sql_query("SELECT COUNT(*) as count FROM branches WHERE is_active=1", conn).iloc[0]['count']
    conn.close()

    st.sidebar.markdown(f"""
    <div style="background: rgba(255,255,255,0.1); padding: 0.75rem; border-radius: 12px; margin: 0.5rem 0;">
        <div style="display: flex; justify-content: space-between; margin-bottom: 0.5rem;">
            <span style="color: white;">📊 Total Members</span>
            <strong style="color: white;">{members_count}</strong>
        </div>
        <div style="display: flex; justify-content: space-between;">
            <span style="color: white;">🏢 Active Branches</span>
            <strong style="color: white;">{branches_count}</strong>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.sidebar.markdown("---")
    if st.sidebar.button("🚪 Logout", use_container_width=True):
        if st.session_state.user_id:
            log_audit(st.session_state.user_id, st.session_state.username, 'LOGOUT', 'user',
                     st.session_state.user_id, "User logged out")
        for key in ['logged_in', 'user_id', 'username', 'user_role', 'user_branch_id', 'branch_name', 'selected_page']:
            if key in st.session_state:
                del st.session_state[key]
        st.rerun()

    st.markdown(f"""
    <div style="text-align: center; margin-bottom: 1rem;">
        <h1 style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                   -webkit-background-clip: text;
                   -webkit-text-fill-color: transparent;
                   margin-bottom: 0;
                   font-size: 2rem;">
            {system_name}
        </h1>
        <p style="color: #666; margin-top: 0.5rem;">Empowering Youth Through Technology & Faith</p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("---")

    if st.session_state.selected_page not in nav_options:
        st.session_state.selected_page = "📝 Registration Form"

    if st.session_state.selected_page == "📝 Registration Form":
        registration_form_ui()
    elif st.session_state.selected_page == "👥 View Members":
        view_members_ui()
    elif st.session_state.selected_page == "📊 Analytics Dashboard":
        analytics_dashboard_ui()
    elif st.session_state.selected_page == "📥 Export Data":
        advanced_export_ui()
    elif st.session_state.selected_page == "📤 Import Data":
        import_data_ui()
    elif st.session_state.selected_page == "🏢 Branch Management":
        branch_management_ui()
    elif st.session_state.selected_page == "👥 User Management":
        user_management_ui()
    elif st.session_state.selected_page == "📧 Email Settings":
        email_settings_ui()
    elif st.session_state.selected_page == "🎨 Theme Customizer":
        theme_customizer_ui()
    elif st.session_state.selected_page == "📜 Audit Log":
        audit_log_ui()

    footer_text = settings.get('footer_text', '© 2024 Kumasi District YoKA. All Rights Reserved.')
    st.sidebar.markdown("---")
    st.sidebar.caption(footer_text)

def main():
    init_rbac_database()
    apply_custom_styling()
    init_session_state()

    if not st.session_state.logged_in:
        login_page()
    else:
        main_dashboard()

if __name__ == "__main__":
    main()
